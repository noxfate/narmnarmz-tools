import Tkinter
import tkFileDialog
import os
import openpyxl
import itertools

def main():
    # fileName = openDialog()
    # openExcelFile("EXM_QM12_Inspection Plan.xlsx")
    newExcel(createFileStructure())

def openDialog():
    root = Tkinter.Tk()
    root.withdraw()
    fileName = tkFileDialog.askopenfilename()
    return fileName

def openExcelFile(filePath):
    wb = openpyxl.load_workbook(filePath)
    sheet_list = wb.get_sheet_names()
    for i in sheet_list:
        print(i)
    type(wb)

def newExcel(sheetList):
    wb = openpyxl.Workbook()
    old_sheet_list = wb.get_sheet_names()
    for i in sheetList:
        wb.create_sheet(title=i)
        sheet = wb.get_sheet_by_name(i)
        for j in range(len(sheetList[i])):
            sheet[openpyxl.utils.get_column_letter(j+1)+'1'] = sheetList[i][j]
    for i in old_sheet_list:
        wb.remove_sheet(wb.get_sheet_by_name(i))    
    wb.save("output_upload.xlsx")

def createFileStructure():
    output_sheets = ["TASK", "MATERIALTASKALLOCATION", "OPERATION"]
    main_header = []
    task_header = ["Key for Task List Group", "Group Counter", "Change Number", "Valid-From Date", "To change number", "Valid-to date", "Deletion Indicator", "Task list usage", "Plant", "Status", "Task list unit of measure", "ISO code for unit of measurement", "From Lot Size", "To Lot Size", "Responsible planner group/department", "Task list description", "Old number", "Recalculate standard values at order creation using CAPP", "Level at Which Dynamic Modification Parameters Are Defined", "Dynamic Modification Rule", "Vendor Considered for Dynamic Modification", "Manufacturer Considered for Dynamic Modification", "Customer Considered for Dynamic Modification", "Sample-Drawing Procedure", "External Numbering of Units to be Inspected", "Identification for the Inspection Point Field Combination", "Partial Lot Assignment in an Inspection During Production", "Change rule", "Change Type for Object", "Work center for capacity planning", "Indicator: Multiple Specifications"]
    material_header = ["Material Number (18 Characters)", "Plant", "Key for Task List Group", "Group Counter", "Change Number", "Valid-From Date", "To change number", "Valid-to date", "Deletion Indicator", "Vendor's account number", "Account number of customer", "Search Field for Customer-Specific Task List Selection", "Long Material Number for MATERIAL Field", "External GUID for MATERIAL Field", "Version Number for MATERIAL Field", "Object for Multiple Specifications", "Type of Object for Multiple Specifications", "Material Number"]
    operation_header = ["Key for Task List Group", "Group Counter", "Change Number", "Valid-From Date", "To change number", "Valid-to date", "Deletion Indicator", "Operation/Activity Number", "Operation ID", "Control key", "Object ID", "Object types of the CIM resource", "Work center", "Plant", "Standard text key", "Operation short text", "Unit of Measure for Activity/Operation", "ISO code for unit of measurement", "Denominator for converting rtg and op units of measure", "Numerator for converting task list and oper. un. of measure", "Base Quantity", "Activity Type", "Unit of measure for the standard value", "ISO code for unit of measurement", "Standard Value", "Activity Type", "Unit of measure for the standard value", "ISO code for unit of measurement", "Standard Value", "Activity Type", "Unit of measure for the standard value", "ISO code for unit of measurement", "Standard Value", "Activity Type", "Unit of measure for the standard value", "ISO code for unit of measurement", "Standard Value", "Activity Type", "Unit of measure for the standard value", "ISO code for unit of measurement", "Standard Value", "Activity Type", "Unit of measure for the standard value", "ISO code for unit of measurement", "Standard Value", "Key word ID for user-defined fields", "User field with 20 characters", "User field with 20 characters", "User field with 10 characters", "User field with 10 characters", "User field for quantity (length 10.3)", "User field: Unit for quantity fields", "ISO code for unit of measurement", "User field for quantity (length 10.3)", "User field: Unit for quantity fields", "ISO code for unit of measurement", "Currency amount for BAPIS (with 9 decimal places)", "User field: Unit for value fields", "ISO currency code", "Currency amount for BAPIS (with 9 decimal places)", "User field: Unit for value fields", "ISO currency code", "User field for date", "User field for date", "User-defined field: Indicator for reports", "User-defined field: Indicator for reports", "Recording View", "Flow Variants for Inspection Point Completion"]
    main_header.append(task_header)
    main_header.append(material_header)
    main_header.append(operation_header)
    result = dict(itertools.izip(output_sheets, main_header))
    return result

    

main()
# output_sheets = ["TASK", "MATERIALTASKALLOCATION", "OPERATION"]
# newExcel(output_sheets)
# a = createFileStructure()
# print(len(a))