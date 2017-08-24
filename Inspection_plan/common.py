import Tkinter
import tkFileDialog
import os
import openpyxl
import itertools
from openpyxl.utils import get_column_letter

def openDialog():
    root = Tkinter.Tk()
    root.withdraw()
    fileName = tkFileDialog.askopenfilename()
    print(fileName)
    return fileName

def openExcelFile(filePath):
    os.chdir(os.path.dirname(filePath))
    wb = openpyxl.load_workbook(os.path.basename(filePath))
    return wb

def findColumnLetterByValueAndRow(worksheet, value, rowNumber):
    for i in range(1, worksheet.max_column+1):
        text = worksheet[get_column_letter(i)+str(rowNumber)].value
        if (text == value):
            return get_column_letter(i)
