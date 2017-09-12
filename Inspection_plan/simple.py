import Tkinter
import tkFileDialog
import os
import openpyxl
import itertools
import easygui
import time
from openpyxl.utils import get_column_letter
from common import *
import sys, traceback

def run():
    filePath = openDialog()
    # wb = openExcelFile(fileName)
    # filePath = "D:/project/narmnarmz-tools/resource/TMP_QM12_Inspection Plan.xlsx"
    start_time = time.time()
    try:
        simple_inspection_structure = configFileStructure()
        data = openExcelFile(filePath)        
        output_filename = composeFileName(filePath)
        newSimpleInspectionPlanExcel(simple_inspection_structure, data, output_filename)
        print("--- %s seconds ---" % (time.time() - start_time))
        easygui.msgbox("Your output is "+output_filename+", which is in the same directory that your selected file. \n\nGood luck, have fun!!\n\nExecutime (s): "+str((time.time() - start_time)), title="Success!")
    except TypeError:
        err = traceback.format_exc()
        easygui.msgbox("TypeError: Maybe this happen because the program can't find field in Excel\n\n"+str(err))
    except UnitConversionError as unit:
        easygui.msgbox("Value Not Found: cannot find this following data in unit.xlsx\n\n"+str(unit))
    except:
        err = traceback.format_exc()
        # print(err)
        easygui.msgbox("Unexpected Error: "+str(err))

def composeFileName(fileFullPath):
    return "UPL_"+os.path.basename(fileFullPath)

def newSimpleInspectionPlanExcel(structure, datamodelwb, fileName):
    wb = openpyxl.Workbook()
    old_sheet_list = wb.get_sheet_names()
    for i in structure:
        wb.create_sheet(title=i)
        sheet = wb.get_sheet_by_name(i)
        for j in range(len(structure[i])):
            sheet[get_column_letter(j+1)+'1'] = structure[i][j]
    for i in old_sheet_list:
        wb.remove_sheet(wb.get_sheet_by_name(i))    

    print("....Start Building....")
    print("....Start Building TASK....")
    buildTaskWorksheet(wb, datamodelwb)
    print("....Start Building TASKALLOCATION....")
    buildMaterialTaskAllocationWorksheet(wb, datamodelwb)
    print("....Start Building OPERATION....")
    buildOperationWorksheet(wb, datamodelwb)
    print("....Start Building INSPCHARACTERISTIC....")
    buildInspcharacteristicWorksheet(wb, datamodelwb)
    print("....Start Building INSP_CHAR_VALUES....")
    buildInspCharValues(wb, datamodelwb)
    print("....Finish Building....")
    print("Output: ", fileName)
    wb.save(fileName)
                
def configFileStructure():
    output_sheets = [
        "TASK", "MATERIALTASKALLOCATION", "OPERATION", "REFERENCEOPERATION", "PRODUCTONRESOURCE", 
        "INSPCHARACTERISTIC", "TEXTALLOCATION", "INSP_CHAR_VALUES"] # index (order) DOES MATTER!!
    main_header = []
    task_header = ["Key for Task List Group", "Group Counter", "Change Number", "Valid-From Date", "To change number", "Valid-to date", "Deletion Indicator", "Task list usage", "Plant", "Status", "Task list unit of measure", "ISO code for unit of measurement", "From Lot Size", "To Lot Size", "Responsible planner group/department", "Task list description", "Old number", "Recalculate standard values at order creation using CAPP", "Level at Which Dynamic Modification Parameters Are Defined", "Dynamic Modification Rule", "Vendor Considered for Dynamic Modification", "Manufacturer Considered for Dynamic Modification", "Customer Considered for Dynamic Modification", "Sample-Drawing Procedure", "External Numbering of Units to be Inspected", "Identification for the Inspection Point Field Combination", "Partial Lot Assignment in an Inspection During Production", "Change rule", "Change Type for Object", "Work center for capacity planning", "Indicator: Multiple Specifications"]
    material_header = ["Material Number (18 Characters)", "Plant", "Key for Task List Group", "Group Counter", "Change Number", "Valid-From Date", "To change number", "Valid-to date", "Deletion Indicator", "Vendor's account number", "Account number of customer", "Search Field for Customer-Specific Task List Selection", "Long Material Number for MATERIAL Field", "External GUID for MATERIAL Field", "Version Number for MATERIAL Field", "Object for Multiple Specifications", "Type of Object for Multiple Specifications", "Material Number"]
    operation_header = ["Key for Task List Group", "Group Counter", "Change Number", "Valid-From Date", "To change number", "Valid-to date", "Deletion Indicator", "Operation/Activity Number", "Operation ID", "Control key", "Object ID", "Object types of the CIM resource", "Work center", "Plant", "Standard text key", "Operation short text", "Unit of Measure for Activity/Operation", "ISO code for unit of measurement", "Denominator for converting rtg and op units of measure", "Numerator for converting task list and oper. un. of measure", "Base Quantity", "Activity Type", "Unit of measure for the standard value", "ISO code for unit of measurement", "Standard Value", "Activity Type", "Unit of measure for the standard value", "ISO code for unit of measurement", "Standard Value", "Activity Type", "Unit of measure for the standard value", "ISO code for unit of measurement", "Standard Value", "Activity Type", "Unit of measure for the standard value", "ISO code for unit of measurement", "Standard Value", "Activity Type", "Unit of measure for the standard value", "ISO code for unit of measurement", "Standard Value", "Activity Type", "Unit of measure for the standard value", "ISO code for unit of measurement", "Standard Value", "Key word ID for user-defined fields", "User field with 20 characters", "User field with 20 characters", "User field with 10 characters", "User field with 10 characters", "User field for quantity (length 10.3)", "User field: Unit for quantity fields", "ISO code for unit of measurement", "User field for quantity (length 10.3)", "User field: Unit for quantity fields", "ISO code for unit of measurement", "Currency amount for BAPIS (with 9 decimal places)", "User field: Unit for value fields", "ISO currency code", "Currency amount for BAPIS (with 9 decimal places)", "User field: Unit for value fields", "ISO currency code", "User field for date", "User field for date", "User-defined field: Indicator for reports", "User-defined field: Indicator for reports", "Recording View", "Flow Variants for Inspection Point Completion"]
    referenceOperation_header = ["Key for Task List Group","Group Counter","Change Number","Valid-From Date","To change number","Valid-to date","Deletion Indicator","Operation ID","Operation/Activity Number","Group of the referenced task list","Group counter of the referenced task list","Increment between referenced operations","Standard text key","Operation short text"]
    productionResource_header = ["Key for Task List Group","Group Counter","Item counter for production resources/tools","Change Number","Valid-From Date","To change number","Valid-to date","Deletion Indicator","Operation/Activity Number","Operation ID","Object types of the CIM resource","Object ID of the resource","Object of a task list","Item Number for Production Resource/Tool","Control key for management of production resources/tools","Indicator: Create load records for prod. resources/tools","Reference date to start of production resource/tool usage","Offset to start of production resource/tool usage","Offset unit for start of prod. resource/tool usage","ISO code for unit of measurement","Reference date for end of production resource/tool usage","Offset to finish of production resource/tool usage","Offset unit for end of production resource/tool usage","ISO code for unit of measurement","Quantity unit of production resource/tool","ISO code for unit of measurement","Standard value for the PRT quantity","Formula for calculating the total quantity of PRT","Usage value unit of the production resource/tool","ISO code for unit of measurement","Standard usage value for production resources/tools","Formula for calculating the total usage value of PRT","Standard text key for production resources/tools","First line of text for production resources/tools","Production resources/tools category","Production resource/tool number","Production resources/tools in plant","Material Number","Production resources and tools","Document Type","Document number","Document Part","Document Version","Equipment Number","Long Material Number for MATERIAL Field","External GUID for MATERIAL Field","Version Number for MATERIAL Field"]
    inspCharacteristic_header = ["Key for Task List Group","Group Counter","Operation ID","Operation/Activity Number","Inspection Characteristic Number","Valid-From Date","Change Number","To change number","Valid-to date","Deletion Indicator","Quantitative Characteristic","Key Containing Preset Characteristic Control Indicators","Master Inspection Characteristics","Plant","Mode for Reference to Master Inspection Characteristic","Short Text for Inspection Characteristic","Inspection Method","Plant for Inspection Method","Tolerance Key","Measured Values Must Be Recorded","Reference to Characteristic Attribute Required","Upper Specification Limit","Lower Specification Limit","Check Target Value","Inspection Scope","Long-Term Inspection","Recording Type","Documentation Required for Inspection Results","Characteristic Category","Sample Quantity Is Added","Destructive Inspection","Calculated Characteristic","Sampling Procedure is Required","Characteristic Relevant for Qual. Score and Scrap Share","Recording the Number of Defects","Assignment of Test Equipment Required","Defects Recording Automatically Called Up","Create Change Documents During Results Recording","SPC Characteristic","Print","Weighting of the Characteristic","Partial Sample No. for Inspection Charac. in Task List","Inspector Qualification","Text Line for Additional Information","Text Line for Additional Information","Text Line for Additional Information","Characteristic Description for Quality Data Exchange","Planned Results Data Origin","Fraction Calculation","Item Number for Production Resource/Tool","Number of Places to the Right of a Decimal Point (Accuracy)","Measurement Unit in Which Quantitative Data Is Maintained","ISO code for unit of measurement","Target Value for a Quantitative Characteristic","Upper Specification Limit","Lower Specification Limit","Number of Value Classes for Inspection Results","Class Width","Class Midpoint","First Upper Specification Limit","First Lower Specification Limit","Second Upper Specification Limit","Second Lower Specification Limit","Upper Plausibility Limit","Lower Plausibility Limit","Check and Calculate Formula in QM","Formula Field","Formula Field","Assigned Code Group or Selected Set","Plant of the Assigned Selected Set","Catalog Type of Assigned Code Group or Selected Set","Assigned Code Group","Catalog Type of Assigned Code Group or Selected Set","Assigned Code Group","Catalog Type of Assigned Code Group or Selected Set","Assigned Code Group","Catalog Type of Assigned Code Group or Selected Set","Assigned Code Group","Defect Code Group for General Rejection","Defect Code for Rejection: General","Defect Code Group for Rejection at Lower Tolerance","Defect Code for Rejection at Lower Specification Limit","Defect Code Group for Rejection at Upper Tolerance","Defect Code for Rejection at Upper Specification Limit","Sampling Procedure in Inspection Characteristic","Sample Unit of Measure","ISO code for unit of measurement","Sample Quantity Factor for Sample(Mult. Sample Unit of Msr.)","SPC Criterion","Dynamic Modification Rule","Characteristic (in Plan) whose Q-Level Will Be Copied","Vendor Considered for Dynamic Modification","Manufacturer Considered for Dynamic Modification","Customer Considered for Dynamic Modification","Parameters for Input Processing in QM Results Recording"]
    textAllocation_header = ["Target type for Direct Input into the EWB","Change Number","Valid-From Date","To change number","Valid-to date","Deletion Indicator","Key for Task List Group","Group Counter","Operation/Activity Number","Operation ID","Item Number for Production Resource/Tool","Item counter for production resources/tools","Inspection Characteristic Number","Single-Character Indicator","Language Key","2-Character SAP Language Code","Text index from","Text index to"]
    inspCharValues_header = ["Key for Task List Group","Group Counter","Operation ID","Operation/Activity Number","Inspection Characteristic Number","Number for Dependent Characteristic Specifications","Change Number","Valid-From Date","To change number","Valid-to date","Material Number","Plant","Vendor Account Number","Account Number of Customer","Search Field for Customer-Specific Task List Selection","Deletion Indicator","Number of Places to the Right of a Decimal Point (Accuracy)","Unit of Measurement, in Which Quantitative Data Is Stored","ISO code for unit of measurement","Target Value for a Quantitative Characteristic","Upper Specification Limit","Lower Specification Limit","Number of Value Classes for Inspection Results","Class Width","Class Midpoint","First Upper Specification Limit","First Lower Specification Limit","Second Upper Specification Limit","Second Lower Specification Limit","Upper Plausibility Limit","Lower Plausibility Limit","Assigned Code Group or Selected Set","Plant of the Assigned Selected Set","Catalog Type of Assigned Code Group or Selected Set","Assigned Code Group","Catalog Type of Assigned Code Group or Selected Set","Assigned Code Group","Catalog Type of Assigned Code Group or Selected Set","Assigned Code Group","Catalog Type of Assigned Code Group or Selected Set","Assigned Code Group","Item Number for Production Resource/Tool","Text Line for Additional Information","Text Line for Additional Information","Text Line for Additional Information","Object for Multiple Specifications","Type of Object for Multiple Specifications","Reference for Standard Specifications"]

    # append order MUST match sheet names in [output_sheets]
    main_header.append(task_header)
    main_header.append(material_header)
    main_header.append(operation_header)
    main_header.append(referenceOperation_header)
    main_header.append(productionResource_header)
    main_header.append(inspCharacteristic_header)
    main_header.append(textAllocation_header)
    main_header.append(inspCharValues_header)
    result = dict(itertools.izip(output_sheets, main_header))
    return result

def buildTaskWorksheet(wb, dataWb):
    ## CONFIG HERE NA N'Narm ##
    DATA_TAB_NAME = "01 - Header" # sheet name to find data
    DATA_ROW_COUNT = 2 # how many row to skip in header
    DATA_HEADER_ROW = 2 # what row to find by field
    ROW_START = 2 # row to start writing data
    IS_FREEZE = True # wanna freeze header ?

    task_ws = wb.get_sheet_by_name("TASK")
    if (IS_FREEZE):
        task_ws.freeze_panes = "A"+ str(ROW_START)
    data_ws = dataWb.get_sheet_by_name(DATA_TAB_NAME)
    n_of_data = data_ws.max_row - DATA_ROW_COUNT
    for i in range(ROW_START, n_of_data + ROW_START): # ROW LOOP in TASK WORKSHEET by n(data)
        for j in range(1, task_ws.max_column+1): #COLUMN LOOP in TASK WORKSHEET by column template
            col = get_column_letter(j)
            if j == 1: # A
                letter = findColumnLetterByColNameAndStartRow(data_ws, "PLNNR", DATA_HEADER_ROW)
                found_data = data_ws[letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 2: # B
                letter = findColumnLetterByColNameAndStartRow(data_ws, "PLNAL", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 4:
                task_ws[col+str(i)] = "01012017"
            elif j == 8: # H
                letter = findColumnLetterByColNameAndStartRow(data_ws, "VERWE", DATA_HEADER_ROW)
                found_data = data_ws[letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 9: # I
                letter = findColumnLetterByColNameAndStartRow(data_ws, "WERKS", DATA_HEADER_ROW)
                found_data = data_ws[letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 10: # J        
                task_ws[col+str(i)] = "4"
            elif j == 11: # K
                letter = findColumnLetterByColNameAndStartRow(data_ws, "MEINH_H", DATA_HEADER_ROW)
                found_data = data_ws[letter+ str(DATA_ROW_COUNT+i-1)].value
                trans_data_cell = transformUnit(found_data)
                trans_data = trans_data_cell.value if trans_data_cell is not None else ""
                task_ws[col+str(i)] = trans_data
            elif j == 13:
                task_ws[col+str(i)] = "1"
            elif j == 14:
                task_ws[col+str(i)] = "99999999"
            elif j == 16:
                letter = findColumnLetterByColNameAndStartRow(data_ws, "KTEXT", DATA_HEADER_ROW)
                found_data = data_ws[letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 24:
                letter = findColumnLetterByColNameAndStartRow(data_ws, "QPRZIEHVER", DATA_HEADER_ROW)
                found_data = data_ws[letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 26:
                letter = findColumnLetterByColNameAndStartRow(data_ws, "SLWBEZ", DATA_HEADER_ROW)
                found_data = data_ws[letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            # else:
            #     task_ws[col+str(i)] = ""
        # print("")
    # return task_ws

def buildMaterialTaskAllocationWorksheet(wb, dataWb):
    ## CONFIG HERE NA N'Narm ##
    DATA_TAB_NAME = "04 - Mat. Assign" # sheet name to find data
    DATA_ROW_COUNT = 2 # how many row to skip in header
    DATA_HEADER_ROW = 2 # what row to find by field
    ROW_START = 2 # row to start writing data
    IS_FREEZE = True # wanna freeze header ?

    task_ws = wb.get_sheet_by_name("MATERIALTASKALLOCATION")
    if (IS_FREEZE):
        task_ws.freeze_panes = "A"+ str(ROW_START)
    data_ws = dataWb.get_sheet_by_name(DATA_TAB_NAME)
    n_of_data = data_ws.max_row - DATA_ROW_COUNT
    for i in range(ROW_START, n_of_data + ROW_START): # ROW LOOP in TASK WORKSHEET by n(data)
        for j in range(1, task_ws.max_column+1): #COLUMN LOOP in TASK WORKSHEET by column template
            col = get_column_letter(j)
            if j == 1: # A
                letter = findColumnLetterByColNameAndStartRow(data_ws, "MATNR", DATA_HEADER_ROW)
                found_data = data_ws[letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 2: # B
                letter = findColumnLetterByColNameAndStartRow(data_ws, "WERKS_A", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 3: 
                letter = findColumnLetterByColNameAndStartRow(data_ws, "PLNNR", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 4:
                letter = findColumnLetterByColNameAndStartRow(data_ws, "PLNAL", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 6:
                task_ws[col+str(i)] = "01012017"
            
def buildOperationWorksheet(wb, dataWb):
    ## CONFIG HERE NA N'Narm ##
    DATA_TAB_NAME = "02 - Operation" # sheet name to find data
    DATA_ROW_COUNT = 2 # how many row to skip in header
    DATA_HEADER_ROW = 2 # what row to find by field
    ROW_START = 2 # row to start writing data
    IS_FREEZE = True # wanna freeze header ?

    task_ws = wb.get_sheet_by_name("OPERATION")
    if (IS_FREEZE):
        task_ws.freeze_panes = "A"+ str(ROW_START)
    data_ws = dataWb.get_sheet_by_name(DATA_TAB_NAME)
    n_of_data = data_ws.max_row - DATA_ROW_COUNT
    for i in range(ROW_START, n_of_data + ROW_START): # ROW LOOP in TASK WORKSHEET by n(data)
        for j in range(1, task_ws.max_column+1): #COLUMN LOOP in TASK WORKSHEET by column template
            col = get_column_letter(j)
            if j == 1: # A
                letter = findColumnLetterByColNameAndStartRow(data_ws, "PLNNR", DATA_HEADER_ROW)
                found_data = data_ws[letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 2: # B
                letter = findColumnLetterByColNameAndStartRow(data_ws, "PLNAL", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 4:
                task_ws[col+str(i)] = "01012017"
            elif j == 8:
                letter = findColumnLetterByColNameAndStartRow(data_ws, "VORNR", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 10:
                task_ws[col+str(i)] = "QM02"
            elif j == 13:
                letter = findColumnLetterByColNameAndStartRow(data_ws, "ARBPL", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 14:
                letter = findColumnLetterByColNameAndStartRow(data_ws, "WERKS", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 16:
                letter = findColumnLetterByColNameAndStartRow(data_ws, "LTXA1", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 19:
                task_ws[col+str(i)] = "1"
            elif j == 21:
                letter = findColumnLetterByColNameAndStartRow(data_ws, "BMSCH", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 23:
                letter = findColumnLetterByColNameAndStartRow(data_ws, "MEINH", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                trans_data_cell = transformUnit(found_data)
                trans_data = trans_data_cell.value if trans_data_cell is not None else ""
                task_ws[col+str(i)] = trans_data

def buildInspcharacteristicWorksheet(wb, dataWb):
    ## CONFIG HERE NA N'Narm ##
    DATA_TAB_NAME = "03 - MIC" # sheet name to find data
    DATA_ROW_COUNT = 3 # how many row to skip in header
    DATA_HEADER_ROW = 3 # what row to find by field
    ROW_START = 2 # row to start writing data
    IS_FREEZE = True # wanna freeze header ?

    task_ws = wb.get_sheet_by_name("INSPCHARACTERISTIC")
    if (IS_FREEZE):
        task_ws.freeze_panes = "A"+ str(ROW_START)
    data_ws = dataWb.get_sheet_by_name(DATA_TAB_NAME)
    n_of_data = data_ws.max_row - DATA_ROW_COUNT
    for i in range(ROW_START, n_of_data + ROW_START): # ROW LOOP in TASK WORKSHEET by n(data)
        for j in range(1, task_ws.max_column+1): #COLUMN LOOP in TASK WORKSHEET by column template
            col = get_column_letter(j)
            if j == 1: # A
                letter = findColumnLetterByColNameAndStartRow(data_ws, "PLNNR", DATA_HEADER_ROW)
                found_data = data_ws[letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 2: # B
                letter = findColumnLetterByColNameAndStartRow(data_ws, "PLNAL", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 4:
                letter = findColumnLetterByColNameAndStartRow(data_ws, "VORNR", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 5:
                letter = findColumnLetterByColNameAndStartRow(data_ws, "MERKNR", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 11:
                letter = findColumnLetterByColNameAndStartRow(data_ws, "QUANTITATIVE_IND", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 13:
                letter = findColumnLetterByColNameAndStartRow(data_ws, "VERWMERKM", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 14:   
                letter = findColumnLetterByColNameAndStartRow(data_ws, "QPMK_WERKS", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 15:
                task_ws[col+str(i)] = "N"
            elif j == 16:   
                letter = findColumnLetterByColNameAndStartRow(data_ws, "QPMK_WERKS", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 17:   
                letter = findColumnLetterByColNameAndStartRow(data_ws, "PMETHODE", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 18:   
                letter = findColumnLetterByColNameAndStartRow(data_ws, "QMTB_WERKS", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 20:   
                letter = findColumnLetterByColNameAndStartRow(data_ws, "MEAS_VALUE_CONFI", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 21:   
                letter = findColumnLetterByColNameAndStartRow(data_ws, "ATTRIBUTE_REQUIR", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 22:   
                letter = findColumnLetterByColNameAndStartRow(data_ws, "UP_TOL_LMT_IND", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 23:   
                letter = findColumnLetterByColNameAndStartRow(data_ws, "LW_TOL_LMT_IND", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 24:   
                letter = findColumnLetterByColNameAndStartRow(data_ws, "TARGET_VAL_CHECK", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 25:   
                letter = findColumnLetterByColNameAndStartRow(data_ws, "SCOPE_IND", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 26:   
                letter = findColumnLetterByColNameAndStartRow(data_ws, "LONG_TERM_INSP_I", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 27:   
                letter = findColumnLetterByColNameAndStartRow(data_ws, "RESULT_RECORDING", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 29:   
                letter = findColumnLetterByColNameAndStartRow(data_ws, "CONFIRMATION_CAT", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 33:
                letter = findColumnLetterByColNameAndStartRow(data_ws, "STICHPRVER", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                if (found_data is not None):                    
                    task_ws[col+str(i)] = "X"
            elif j == 39:
                letter = findColumnLetterByColNameAndStartRow(data_ws, "SPC_CRITERION_KEY", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                if (found_data is not None):                    
                    task_ws[col+str(i)] = "X"
            elif j == 40:   
                letter = findColumnLetterByColNameAndStartRow(data_ws, "PRINT_IND", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 43:   
                letter = findColumnLetterByColNameAndStartRow(data_ws, "INSPECTOR_QUALIF", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 51:   
                letter = findColumnLetterByColNameAndStartRow(data_ws, "DEC_PLACES", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 52:   
                letter = findColumnLetterByColNameAndStartRow(data_ws, "MEAS_UNIT", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                trans_data_cell = transformUnit(found_data)
                trans_data = trans_data_cell.value if trans_data_cell is not None else ""
                task_ws[col+str(i)] = trans_data
            elif j == 54:   
                letter = findColumnLetterByColNameAndStartRow(data_ws, "SOLLWERT", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 55:   
                letter = findColumnLetterByColNameAndStartRow(data_ws, "TOLERANZOB", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 56:   
                letter = findColumnLetterByColNameAndStartRow(data_ws, "TOLERANZUN", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 66:   
                letter = findColumnLetterByColNameAndStartRow(data_ws, "FORMULA_IND", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 67:   
                letter = findColumnLetterByColNameAndStartRow(data_ws, "FORMEL1", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 68:   
                letter = findColumnLetterByColNameAndStartRow(data_ws, "FORMEL2", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 69:   
                letter = findColumnLetterByColNameAndStartRow(data_ws, "AUSWMENGE1", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 70:   
                letter = findColumnLetterByColNameAndStartRow(data_ws, "QWERKAUSW", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 71:   
                letter = findColumnLetterByColNameAndStartRow(data_ws, "AUSWMENGE1", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                if (found_data is not None):                    
                    task_ws[col+str(i)] = "1"    
            elif j == 85:   
                letter = findColumnLetterByColNameAndStartRow(data_ws, "STICHPRVER", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data  
            elif j == 86:   
                letter = findColumnLetterByColNameAndStartRow(data_ws, "PROBEMGEH", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                trans_data_cell = transformUnit(found_data)
                trans_data = trans_data_cell.value if trans_data_cell is not None else ""
                task_ws[col+str(i)] = trans_data
            elif j == 88:   
                letter = findColumnLetterByColNameAndStartRow(data_ws, "STICHPRVER", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                if (found_data is not None):                    
                    task_ws[col+str(i)] = "1"   
            elif j == 89:   
                letter = findColumnLetterByColNameAndStartRow(data_ws, "SPC_CRITERION_KEY", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data         

def buildInspCharValues(wb, dataWb):
    ## CONFIG HERE NA N'Narm ##
    DATA_TAB_NAME = "05 - Denp. Char." # sheet name to find data
    DATA_ROW_COUNT = 2 # how many row to skip in header
    DATA_HEADER_ROW = 2 # what row to find by field
    ROW_START = 2 # row to start writing data
    IS_FREEZE = True # wanna freeze header ?

    task_ws = wb.get_sheet_by_name("INSP_CHAR_VALUES")
    if (IS_FREEZE):
        task_ws.freeze_panes = "A"+ str(ROW_START)
    data_ws = dataWb.get_sheet_by_name(DATA_TAB_NAME)
    n_of_data = data_ws.max_row - DATA_ROW_COUNT
    for i in range(ROW_START, n_of_data + ROW_START): # ROW LOOP in TASK WORKSHEET by n(data)
        for j in range(1, task_ws.max_column+1): #COLUMN LOOP in TASK WORKSHEET by column template
            col = get_column_letter(j)
            if j == 1: # A
                letter = findColumnLetterByColNameAndStartRow(data_ws, "PLNNR", DATA_HEADER_ROW)
                found_data = data_ws[letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 2: # B
                letter = findColumnLetterByColNameAndStartRow(data_ws, "PLNAL", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 4:
                letter = findColumnLetterByColNameAndStartRow(data_ws, "VORNR", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 5:
                letter = findColumnLetterByColNameAndStartRow(data_ws, "MERKNR", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 6:
                letter = findColumnLetterByColNameAndStartRow(data_ws, "ZUORDNR", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 11:
                letter = findColumnLetterByColNameAndStartRow(data_ws, "MATNR", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 12:
                letter = findColumnLetterByColNameAndStartRow(data_ws, "WERKS_A", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 13:
                letter = findColumnLetterByColNameAndStartRow(data_ws, "LIFNR", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 17:
                letter = findColumnLetterByColNameAndStartRow(data_ws, "DEC_PLACES", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 18:
                letter = findColumnLetterByColNameAndStartRow(data_ws, "MEAS_UNIT", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                trans_data_cell = transformUnit(found_data)
                trans_data = trans_data_cell.value if trans_data_cell is not None else ""
                task_ws[col+str(i)] = trans_data
            elif j == 20:
                letter = findColumnLetterByColNameAndStartRow(data_ws, "SOLLWERT", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 21:
                letter = findColumnLetterByColNameAndStartRow(data_ws, "TOLERANZOB", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 22:
                letter = findColumnLetterByColNameAndStartRow(data_ws, "TOLERANZUN", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 32:
                letter = findColumnLetterByColNameAndStartRow(data_ws, "AUSWMENGE1", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 33:
                letter = findColumnLetterByColNameAndStartRow(data_ws, "QWERKAUSW", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                task_ws[col+str(i)] = found_data
            elif j == 34:
                letter = findColumnLetterByColNameAndStartRow(data_ws, "AUSWMENGE1", DATA_HEADER_ROW)
                found_data = data_ws[ letter+ str(DATA_ROW_COUNT+i-1)].value
                if (found_data is not None):                    
                    task_ws[col+str(i)] = "1"
