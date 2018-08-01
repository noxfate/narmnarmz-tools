import Tkinter
import tkFileDialog
import os
import openpyxl
import itertools
import easygui
import time
from openpyxl.utils import get_column_letter
from common import *
import sys, traceback

def enum(**enums):
    return type('Enum', (), enums)

ValidateError = enum(
        NOT_NULL=["Not null", "{} cannot be null"],
        NULL=["Null", "Leave {} blank"],
        VALUE_TYPE=["Value type", "{} is incorrect data"],
        LENGTH=["Length", "{} is out of length"],
        FIXED_VALUE=["Fixed value", "{} must be {}"],
        FIXED_VALUE_EMPTY=["Fixed value Not Found in Dict", "{} doesn't exist"],
        FIXED_VALUE_X=["Fixed value X", "X must be capital letter"],
        DUPLICATE_KEY=["Duplicate", "Duplicate code"],
        DUPLICATE=["Duplicate", "Duplicate material assignment"],
        UNDEFINED=["Undefined", "{}"]
    )

cur_path = os.path.dirname(__file__)
new_unit_path = os.path.join(cur_path,'..', 'resource','Dict', 'unit.xlsx')
#new_val_dict_path = os.path.join(cur_path, '..','resource','Dict', '02 Dictionary V1.0.XLSX')
#val_dict_wb = openExcelFile(new_val_dict_path)

def find_in_dict(sheetName, colNumber, input):
    global val_dict_wb
    ws = val_dict_wb.get_sheet_by_name(sheetName)
    if input is not None:
        found = findCellInColumnByValue(ws, colNumber, input, 0)
        # print("Find: ["+sheetName+"] "+str(input), ", Found: "+str(found))
        if found is None:
            return None
        return found
    return None

def find_multiple_in_dict(sheetName, inputDict):
    global val_dict_wb
    ws = val_dict_wb.get_sheet_by_name(sheetName)
    found = []
    if inputDict is not None:
        for k in inputDict.keys():
            cells = findCellListInColumnByValue(ws, k, inputDict[k], 1)
            rows = []
            if cells is not None:
                cells_list = list(cells)
                for i in range(len(cells_list)):
                    rows.append(cells_list[i].row)
            found.append(set(rows))
        result = set.intersection(*found)
        return result
    return set()


import validate_01header
import validate_02operation
import validate_03mic
import validate_04mat
#import validate_05denp

#varSheet = select sheet to validate
#varB = Business (factory, farm)
#varAdd = run Additional condition or not (MIC only)

def run(varSheet,varB,varAdd):
    filePath = openDialog()
    # filePath = "D:/project/narmnarmz-tools/resource/TMP_QM12_Inspection Plan.xlsx"
    start_time = time.time()

    if varB == "Factory":
        new_val_dict_path = os.path.join(cur_path, '..','resource','Dict', '02 Dictionary V1.0.XLSX')
    else:
        new_val_dict_path = os.path.join(cur_path, '..','resource','Dict', '02 Dictionary Farm.xlsx')

    global val_dict_wb
    val_dict_wb = openExcelFile(new_val_dict_path)

    try:
        file_structure = configFileStructure()
        data = openExcelFile(filePath)        
        output_filename = composeFileName(filePath,varSheet)
        newValidateInspExcel(file_structure, data, output_filename,varSheet ,varB, varAdd)
        print("--- %s seconds ---" % (time.time() - start_time))
        easygui.msgbox("Your output is "+output_filename+", which is in the same directory that your selected file. \n\nGood luck, have fun!!\n\nExecutime (s): "+str((time.time() - start_time)), title="Success!")
    except TypeError:
        err = traceback.format_exc()
        print(err)
        easygui.msgbox("TypeError: Maybe this happen because the program can't find field in Excel\n\n"+str(err))
    except UnitConversionError as dic:
        easygui.msgbox("Value Not Found: cannot find this following data in 02 Dictionary V1.0.xlsx\n\n"+str(dic))
    except:
        err = traceback.format_exc()
        print(err)
        easygui.msgbox("Unexpected Error: "+str(err))


def composeFileName(fileFullPath, varSheet):
    sheet = ""
    for item in varSheet:
        sheet = sheet+str(item)
    return "ERR_"+sheet+"_"+os.path.basename(fileFullPath)

def newValidateInspExcel(structure, datamodelwb, fileName,varSheet, varB, varAdd):
    wb = openpyxl.Workbook(write_only=True)
    old_sheet_list = wb.get_sheet_names()
    for i in structure:
        wb.create_sheet(title=i)
        sheet = wb.get_sheet_by_name(i)
        for j in range(len(structure[i])):
            sheet[get_column_letter(j+1)+'1'] = structure[i][j]
    for i in old_sheet_list:
        wb.remove_sheet(wb.get_sheet_by_name(i))    

    print("....Start Building....")

    for item in varSheet:
        
        if item==1:
            print("....Validating 01 - Header....")
            validate_01header.validate(wb, datamodelwb, varB)
            #print varSheet[i-1]
        if item==2:
            print("....Validating 02 - Operaion....")
            validate_02operation.validate(wb, datamodelwb,varB)
            #print varSheet[i-1]
        if item==3:
            print("....Validating 03 - MIC....")
            validate_03mic.validate(wb, datamodelwb, varB ,varAdd)
            #print varSheet[i-1]
        if item==4:
            print("....Validating 04 - Mat. Assign....")
            validate_04mat.validate(wb, datamodelwb, varB)
            #print varSheet[i-1]
        #if item==5:
            #print("....Validating 05 - Denp. Char. ....")
            #validate_05denp.validate(wb, datamodelwb)
            #print varSheet[i-1]
        
    print("Output: ", fileName)
    wb.save(fileName)

def configFileStructure():
    output_sheets = [
        "01 - Header", "02 - Operation", "03 - MIC", "04 - Mat. Assign", "05 - Denp. Char."
        ] # index (order) DOES MATTER!!
    main_header = []
    header_01 = ["Status", "Group", "Group Counter", "Plant", "Task list description", "Error Message"]
    header_02 = ["Status", "Group", "Group Counter", "Operation/Activity", "Operation short text", "Error Message"]
    header_03 = ["Status", "Group", "Group Counter", "Operation/Activity", "Characteristic number", "Error Message"]
    header_04 = ["Status", "Group", "Group Counter", "Assign Plant", "Material", "Error Message"]
    #header_05 = ["Status", "Group", "Group Counter", "Operation Number", "Characteristic Numner", "No. Dep. Char. Specs", "Error Message"]

    # append order MUST match sheet names in [output_sheets]
    main_header.append(header_01)
    main_header.append(header_02)
    main_header.append(header_03)
    main_header.append(header_04)
    #main_header.append(header_05)
    result = dict(itertools.izip(output_sheets, main_header))
    return result