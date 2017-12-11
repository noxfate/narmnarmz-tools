from common import *
from openpyxl.utils import get_column_letter
import openpyxl
from validate import ValidateError, find_in_dict, find_multiple_in_dict

def writeHeaderReport(ws, status, data, errorMsg, debug=None, colName=None, isQL=None):
    new_row = []
    new_row.append(status)
    if len(data) != 4:
        raise ValueError("[03-MIC] Data size is not correct")
    new_row += data
    new_row.append(errorMsg)
    new_row.append(debug)
    new_row.append(colName)
    new_row.append(isQL)
    insert_new_row(ws, new_row) 

def get_01header_SLWBEZ_by_key(dataWb, keyDict):
    ws = dataWb.get_sheet_by_name("01 - Header")
    keys = dict(keyDict)
    keys.pop("VORNR")
    keys.pop("MERKNR")
    found = find_by_keys(ws, 2, 2, keys)
    found = list(found)
    found.sort()
    row = found[0] if len(found) >= 1 else 0
    if row == 0:
        return False
    col = findColumnLetterByColNameAndStartRow(ws, "SLWBEZ", 2)
    SLWBEZ = ws[col+str(row)].value
    return SLWBEZ

def get_value_by_row_colname(ws, colname, row):
    MIC_HEADER = 3
    col = findColumnLetterByColNameAndStartRow(ws, colname, MIC_HEADER)
    if ws[col+str(row)].value == '':
        return None
    return ws[col + str(row)].value

def check_same_matassign_by_MEINS(dataWb, keyDict, data):
    ws = dataWb.get_sheet_by_name("04 - Mat. Assign")
    keys = dict(keyDict)
    keys.pop("VORNR")
    keys.pop("MERKNR")
    found = find_by_keys(ws, 2, 2, keys)
    found = list(found)
    found.sort()
    row = found[0] if len(found) >= 1 else 0
    if row == 0:
        return True
    col = findColumnLetterByColNameAndStartRow(ws, "MEINS", 2)
    WERKS = ws[col + str(row)].value
    return data == WERKS

def get_decimal_places(ws, row):
    MIC_HEADER = 3
    col = findColumnLetterByColNameAndStartRow(ws, "DEC_PLACES", MIC_HEADER)
    dec = ws[col + str(row)].value
    if dec is None or (isinstance(dec, str) and dec.strip() == '') or (isinstance(dec, unicode) and dec.strip() == '') or not isNumOnly(dec):
        return 0
    return dec


def validate(wb, dataWb):

    ## CONFIG HERE NA N'Narm ##
    DATA_TAB_NAME = "03 - MIC" # sheet name to find data
    DATA_ROW_COUNT = 3 # how many row to skip in header
    DATA_FIELD_ROW = 1
    DATA_HEADER_ROW = 3 # what row to find by field
    ROW_START = 2 # row to start writing data
    IS_FREEZE = True # wanna freeze header ?
    active_ws = wb.get_sheet_by_name("03 - MIC")
    opr_ws = dataWb.get_sheet_by_name("02 - Operation")

    if (IS_FREEZE):
        active_ws.freeze_panes = "A"+ str(ROW_START)
    data_ws = dataWb.get_sheet_by_name(DATA_TAB_NAME)
    n_of_data = data_ws.max_row - DATA_ROW_COUNT

    # CHECK Addtional Condition 1-2
    PLNNR_col = findColumnLetterByColNameAndStartRow(data_ws, "PLNNR", DATA_HEADER_ROW)
    PLNAL_col = findColumnLetterByColNameAndStartRow(data_ws, "PLNAL", DATA_HEADER_ROW)
    VORNR_col = findColumnLetterByColNameAndStartRow(data_ws, "VORNR", DATA_HEADER_ROW)
    MERKNR_col = findColumnLetterByColNameAndStartRow(data_ws, "MERKNR", DATA_HEADER_ROW)

    
    for i in range(DATA_ROW_COUNT+1, n_of_data + DATA_ROW_COUNT+1):

        PLNNR = data_ws[PLNNR_col + str(i)].value
        PLNAL = data_ws[PLNAL_col + str(i)].value
        VORNR = data_ws[VORNR_col + str(i)].value
        MERKNR = data_ws[MERKNR_col + str(i)].value

        data = [PLNNR, PLNAL, VORNR, MERKNR]
        if (PLNNR is None or PLNAL is None or VORNR is None or MERKNR is None or len(str(VORNR))!=4 or len(str(MERKNR))!=4):
            #writeHeaderReport(active_ws, "WARNING", data, ValidateError.UNDEFINED[1].format("Some keys are null and will be skip"), "row="+str(i))
            continue

        d = dict()
        d["PLNNR"] = PLNNR
        d["PLNAL"] = PLNAL
        d["VORNR"] = VORNR
        d["MERKNR"] = MERKNR
        match_cond_1 = find_by_keys(data_ws, DATA_HEADER_ROW, DATA_ROW_COUNT, d)
        #cond_1 = check_duplicate_key(data_ws, DATA_HEADER_ROW, DATA_ROW_COUNT, d)
        # print("Cond1", match_cond_1)

        #opr_ws = dataWb.get_sheet_by_name("02 - Operation")
        d = dict()
        d["PLNNR"] = PLNNR
        d["PLNAL"] = PLNAL
        d["VORNR"] = VORNR
        match_cond_2 = find_by_keys(opr_ws, 2, 2, d)
        #cond_2 = is_key_exist(opr_ws, 2, 2, d)
        # print("Cond2", match_cond_2)
		
        if len(match_cond_1) > 1:
            writeHeaderReport(active_ws, "ERROR", data, ValidateError.DUPLICATE_KEY[1], "N="+str(len(match_cond_1)))
        if len(match_cond_2) < 1:
            writeHeaderReport(active_ws, "ERROR", data, ValidateError.UNDEFINED[1].format("Group does not exist in 02 - Operation"), "N="+str(len(match_cond_2)))
        #if cond_1:
            #writeHeaderReport(active_ws, "ERROR", data, ValidateError.DUPLICATE_KEY[1], "row="+str(i))

        #if not cond_2:
            #writeHeaderReport(active_ws, "ERROR", data, ValidateError.UNDEFINED[1].format("Group does not exist in 02 - Operation"), "row="+str(i))
    
    print("Fin Additional Condition")
	
    # Check By Field
    key = ["PLNNR", "PLNAL", "VORNR", "MERKNR"]
    for i in range(DATA_ROW_COUNT+1, n_of_data + DATA_ROW_COUNT+1):
    	
    	progress = (float(format(i,'.4f'))/n_of_data)*100
    	if i > 1000 and i%1000 == 0:
    		print "....Validating 03 - MIC " + str(format(progress,'.2f') +" % (" + str(i) + "/" + str(n_of_data) + ")")

        report_data = [
                data_ws[PLNNR_col+str(i)].value, 
                data_ws[PLNAL_col+str(i)].value,
                data_ws[VORNR_col+str(i)].value,
                data_ws[MERKNR_col+str(i)].value,
            ]
        key_value = [
                data_ws[PLNNR_col+str(i)].value, 
                data_ws[PLNAL_col+str(i)].value,
                data_ws[VORNR_col+str(i)].value,
                data_ws[MERKNR_col+str(i)].value,
            ]
        key_data_dict = dict(itertools.izip(key,key_value))

        QPMK_WERKS_col = findColumnLetterByColNameAndStartRow(data_ws, "QPMK_WERKS", DATA_HEADER_ROW)
        VERKMERKM_col = findColumnLetterByColNameAndStartRow(data_ws, "VERWMERKM", DATA_HEADER_ROW)
        QPMK_WERKS = data_ws[QPMK_WERKS_col + str(i)].value
        VERWMERKM = data_ws[VERKMERKM_col + str(i)].value
        if QPMK_WERKS is None:
            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format("Plant - MIC"), i)
        if VERWMERKM is None:
            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format("Master Inspection Characteristic"), i)

        x_dict = dict()
        x_dict[2] = QPMK_WERKS
        x_dict[3] = VERWMERKM
        found_QL = find_multiple_in_dict("06-MICQL", x_dict)
        found_QN = find_multiple_in_dict("06-MICQN", x_dict)
        if len(found_QL) == 0 and len(found_QN) != 0:
            isQL = False
        elif len(found_QL) != 0 and len(found_QN) == 0:
            isQL = True
        elif len(found_QL) == 0 and len(found_QN) == 0:
            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Doesn't exist"), i)
            continue
        #else:
            #writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Found in both MICQL, MICQN"), i)
            #continue

        for j in range(1, data_ws.max_column +1):
            field_descr = data_ws.cell(row=DATA_FIELD_ROW, column=j).value
            real_data = data_ws.cell(row=i, column=j).value
            if isinstance(real_data, int) or isinstance(real_data, long) or isinstance(real_data, float):
                data = str(real_data)
            else:
                data = real_data
            
            if data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "PLNNR":                
                if isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                elif not isNull(data) and len(data) > 15:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "PLNAL":
                if isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                elif not isNumeric(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                if not isNull(data) and len(data) > 2:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "VORNR":
                if isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                elif not isNull(data) and len(data) != 4:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Operation should have 4 digits"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                elif not isNumOnly(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "MERKNR":
                if isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                if not isNull(data) and len(data) != 4:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Characteristic number should have 4 digits"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                if not isNumOnly(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "QUANTITATIVE_IND":
                if isQL:
                    if not isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Qualitative: QUANTITATIVE_IND must be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                else:
                    if isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    if not isNull(data) and data != "X":
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "X"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "MKVERSION":
                if not isNull(data) and data != "1":
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "1"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "KURZTEXT":
                if not isNull(data) and len(data) > 40:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "ATTRIBUTE_REQUIR":
                if not isQL:
                    if not isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Quantitative: ATTRIBUTE_REQUIR must be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                else:
                    if isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    if not isNull(data) and data != "X":
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "X"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "MEAS_VALUE_CONFI":
                if isQL:
                    if not isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Qualitative: MEAS_VALUE_CONFI must be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                else:
                    if isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    if not isNull(data) and data != "X":
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "X"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "RESULT_RECORDING":
                if not isNull(data) and data != "+":
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "+"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "AUTO_DEFCT_RECOR":
                if not isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "CONFIRMATION_CAT":
                if not isNull(data) and data != "X":
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "X"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "ADD_SAMPLE_QUANT":
                if not isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "DESTRUCTIVE_INSP":
                if not isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "SPC_IND":
                if isQL:
                    if not isNull(data):
                         writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Qualitative: SPC_IND must be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                else:                    
                    SPC_CRIT = get_value_by_row_colname(data_ws, "SPC_CRITERION_KEY", i)
                    if isNull(SPC_CRIT) and not isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("SPC_IND Conflict with SPC_CRITERION_KEY"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    elif SPC_CRIT is not None and data != "X":
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "X"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "SPC_CRITERION_KEY":
                if isQL:
                    if not isNull(data):
                         writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Qualitative: SPC_CRITERION_KEY must be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                else:
                    SPC_IND = get_value_by_row_colname(data_ws, "SPC_IND", i)
                    if isNull(SPC_IND) and not isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("SPC_IND Conflict with SPC_CRITERION_KEY"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    elif SPC_IND is not None and data != "070":
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "070"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "SAMPLING_PROCEDU":
                if isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                if not isNull(data) and data != "X":
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "X"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "STICHPRVER":
                if isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                if not isNull(data) and len(data) > 8:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                elif not isNull(data):
                    Header_SLWBEZ = get_01header_SLWBEZ_by_key(dataWb, key_data_dict)
                    MIC_SPC = get_value_by_row_colname(data_ws, "SPC_IND", i)
                    found = None
                    if str(VERWMERKM)[0] == 'F':
                        if data != 'FFF0NLAB':
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Incorrect sampling procedure"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    elif isNull(Header_SLWBEZ) and isNull(MIC_SPC):
                        found = find_in_dict("07-Samp", 1, real_data)
                    elif isNull(Header_SLWBEZ) and not isNull(MIC_SPC):
                        found = find_in_dict("07-SampSPC", 1, real_data)
                    elif not isNull(Header_SLWBEZ) and isNull(MIC_SPC):
                        found = find_in_dict("07-SampPoint", 1, real_data)
                    elif not isNull(Header_SLWBEZ) and not isNull(MIC_SPC):
                        found = find_in_dict("07-SampPointSPC", 1, real_data)
                    if isNull(found):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Incorrect sampling procedure"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "PROBEMGEH":
                if isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                if not isNull(data) and len(data) > 6:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                if not check_same_matassign_by_MEINS(dataWb, key_data_dict, real_data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Sampling Unit of Measure not mapping with Material Master"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "PRUEFEINH":
                if isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                if not isNull(data) and not isNumeric(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                if not isNull(data) and data != "1":
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "1"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "MEAS_UNIT":
                if isQL:
                    if not isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Qualitative: MEAS_UNIT must be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                else:
                    if not isNull(data) and len(data) > 6:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    if not isNull(data) and find_in_dict("02-Unit", 4, real_data) is None:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MEAS_UNIT doesn't exist"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "TARGET_VAL_CHECK":
                if isQL:
                    if not isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Qualitative: TARGET_VAL_CHECK must be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                else:
                    if get_value_by_row_colname(data_ws, "SOLLWERT", i) is not None:
                        if isNull(data):
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                        if not isNull(data) and data != "X":                    
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Conflict with Target Value"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "LW_TOL_LMT_IND":
                if isQL:
                    if not isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Qualitative: LW_TOL_LMT_IND must be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                else:
                    if get_value_by_row_colname(data_ws, "TOLERANZUN", i) is not None:
                        if isNull(data):
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                        if not isNull(data) and data != "X":                    
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Conflict with Lower Limit"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "UP_TOL_LMT_IND":
                if isQL:
                    if not isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Qualitative: UP_TOL_LMT_IND must be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                else:
                    if get_value_by_row_colname(data_ws, "TOLERANZOB", i) is not None:
                        if isNull(data):
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                        if not isNull(data) and data != "X":                    
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Conflict with Upper Limit"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "DEC_PLACES":
                if isQL:
                    if not isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Qualitative: DEC_PLACES must be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                else:
                    if not isNull(data) and not isNumeric(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    if not isNull(data) and len(data) > 3:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "SOLLWERT":
                if isQL:
                    if not isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Qualitative: Target Value must be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                else:
                    if not isNull(data) and not isNumeric(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    elif not isNull(data) and len(data) > 16:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    elif get_value_by_row_colname(data_ws, "TARGET_VAL_CHECK", i) is not None:
                        if isNull(data):
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Conflict with TARGET_VAL_CHECK"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                        dec = get_decimal_places(data_ws, i)
                        if not checkDecimalPlace(dec, data):
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Target Value conflict with Decimal place"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                        if not isNumeric(data):
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                        v1 = get_value_by_row_colname(data_ws, "TOLERANZUN", i) # lower
                        v2 = get_value_by_row_colname(data_ws, "TOLERANZOB", i) # upper
                        if not isNull(real_data):
                            try:
                                if v1 is not None:
                                    if float(real_data) < float(v1):
                                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Target Value conflict with Lower Limit"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                                if v2 is not None:
                                    if float(real_data) > float(v2):
                                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Target Value conflict with Upper Limit"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                            except ValueError:
                                writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format("SOLLWERT, TOLERANZUN, TOLERANZOB"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)

            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "TOLERANZUN":
                if isQL:
                    if not isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Qualitative: Lower Limit must be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                else:
                    if not isNull(data) and not isNumeric(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    elif not isNull(data) and len(data) > 16:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    elif get_value_by_row_colname(data_ws, "LW_TOL_LMT_IND", i) is not None:
                        if isNull(data):
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Conflict with LW_TOL_LMT_IND"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                        if not isNumeric(data):
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                        dec = get_decimal_places(data_ws, i)
                        if not checkDecimalPlace(dec, data):
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Lower Limit conflict with Decimal place"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                        v1 = get_value_by_row_colname(data_ws, "SOLLWERT", i) # target
                        v2 = get_value_by_row_colname(data_ws, "TOLERANZOB", i) # upper
                        if not isNull(real_data):    
                            try:
                                if v1 is not None:
                                    if float(real_data )> float(v1):
                                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Lower Limit conflict with Target Value"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                                if v2 is not None:
                                    if float(real_data )> float(v2):
                                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Lower Limit conflict with Upper Value"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                            except ValueError:
                                writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format("SOLLWERT, TOLERANZUN, TOLERANZOB"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "TOLERANZOB":
                if isQL:
                    if not isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Qualitative: Upper Limit must be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                else:
                    if not isNull(data) and not isNumeric(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    elif not isNull(data) and len(data) > 16:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    elif get_value_by_row_colname(data_ws, "UP_TOL_LMT_IND", i) is not None:
                        if isNull(data):
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Conflict with UP_TOL_LMT_IND"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                        if not isNumeric(data):
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                        dec = get_decimal_places(data_ws, i)
                        if not checkDecimalPlace(dec, data):
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Upper Limit conflict with Decimal place"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                        v1 = get_value_by_row_colname(data_ws, "SOLLWERT", i) # target
                        v2 = get_value_by_row_colname(data_ws, "TOLERANZUN", i) # lower
                        if not isNull(real_data):
                            try:
                                if v1 is not None:
                                    if float(real_data )< float(v1):
                                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Upper Limit conflict with Target Value"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                                if v2 is not None:
                                    if float(real_data )< float(v2):
                                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Upper Limit conflict with Lower Limit"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                            except ValueError:
                                writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format("SOLLWERT, TOLERANZUN, TOLERANZOB"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "AUSWMENGE1":
                if isQL:
                    if isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Selected set cannot be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    if not isNull(data) and len(data) > 8:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    QWERKAUSW = get_value_by_row_colname(data_ws, "QWERKAUSW", i)
                    x_dict = dict()
                    x_dict[3] = real_data
                    x_dict[2] = QWERKAUSW
                    if find_multiple_in_dict("08-Selected Set", x_dict) is None:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE_EMPTY[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                else:
                    if not isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Quantitative: Selected set must be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "QWERKAUSW":
                if isQL:
                    if isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Plant for Selected set cannot be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    if not isNull(data) and len(data) > 4:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    AUSWMENGE1 = get_value_by_row_colname(data_ws, "AUSWMENGE1", i)
                    x_dict = dict()
                    x_dict[3] = AUSWMENGE1
                    x_dict[2] = real_data
                    if find_multiple_in_dict("08-Selected Set", x_dict) is None:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE_EMPTY[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                else:
                    if not isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Quantitative: Plant for Selected set must be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "PMETHODE":
                if not isNull(data) and len(data) > 8:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                if not isNull(data) and find_in_dict("09-Method",3, real_data) is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE_EMPTY[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "QMTB_WERKS":
                if not isNull(data) and get_value_by_row_colname(data_ws, "PMETHODE", i) is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("PMETHODE conflict with QMTB_WERKS"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                if not isNull(data) and len(data) > 4:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                if not isNull(data) and find_in_dict("09-Method",2, real_data) is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE_EMPTY[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "PMTVERSION":
                if not isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "LONG_TERM_INSP_I":
                if not isNull(data) and data != "X":
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "X"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "INSPECTOR_QUALIF":
                if not isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "DUMMY10":
                if not isNull(data) and len(data) > 10:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "DUMMY20":
                if not isNull(data) and len(data) > 20:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "DUMMY40":
                if not isNull(data) and len(data) > 40:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "SCOPE_IND":
                if get_value_by_row_colname(data_ws, "SPC_IND", i) is not None:
                    if not isNull(data) and data != "=":
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "="), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "QSCORE_AND_SHARE":
                if not isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "CHANGE_DOCUMENTS":
                if not isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "INSP_TOOL_IND":
                if not isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "DOCU_REQU":
                if not isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "PRINT_IND":
                if not isNull(data) and data != "X":
                    writeHeaderReport(active_ws, "ERROR", report_data,  ValidateError.FIXED_VALUE[1].format(field_descr, "X"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "FORMULA_IND":
                if isQL:
                    if not isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Qualitative: FORMULA_IND must be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                else:
                    if get_value_by_row_colname(data_ws, "FORMEL1", i) is not None:
                        if isNull(data) or data != "X":
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "X"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "FORMEL1":
                if isQL:
                    if not isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Qualitative: FORMULA_FIELD_1 must be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                else:
                    if get_value_by_row_colname(data_ws, "FORMULA_IND", i) is not None:
                        if isNull(data):
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("FORMEL1 conflict with FORMULA_IND"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                        if not isNull(data) and len(data) > 60:
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "FORMEL2":
                if isQL:
                    if not isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Qualitative: FORMULA_FIELD_2 must be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                else:
                    if get_value_by_row_colname(data_ws, "FORMEL2", i) is not None:
                        if isNull(data):
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Fill formula in FORMEL1 first"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                        if not isNull(data) and len(data) > 60:
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)