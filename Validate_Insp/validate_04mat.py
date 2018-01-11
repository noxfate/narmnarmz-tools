from common import *
from openpyxl.utils import get_column_letter
import openpyxl
from validate import ValidateError, find_in_dict

def writeHeaderReport(ws, status, data, errorMsg, debug=None):
    new_row = []
    new_row.append(status)
    if len(data) != 4:
        raise ValueError("[04-Mat. Assign] Data size is not correct")
    new_row += data
    new_row.append(errorMsg)
    new_row.append(debug)
    insert_new_row(ws, new_row)  

def validate(wb, dataWb, varB):
    ## CONFIG HERE NA N'Narm ##
    DATA_TAB_NAME = "04 - Mat. Assign" # sheet name to find data
    DATA_ROW_COUNT = 2 # how many row to skip in header
    DATA_FIELD_ROW = 1
    DATA_HEADER_ROW = 2 # what row to find by field
    ROW_START = 2 # row to start writing data
    IS_FREEZE = True # wanna freeze header ?

    active_ws = wb.get_sheet_by_name("04 - Mat. Assign")
    if (IS_FREEZE):
        active_ws.freeze_panes = "A"+ str(ROW_START)
    data_ws = dataWb.get_sheet_by_name(DATA_TAB_NAME)
    n_of_data = data_ws.max_row - DATA_ROW_COUNT

    # CHECK Addtional Condition 1-2
    PLNNR_col = findColumnLetterByColNameAndStartRow(data_ws, "PLNNR", DATA_HEADER_ROW)
    PLNAL_col = findColumnLetterByColNameAndStartRow(data_ws, "PLNAL", DATA_HEADER_ROW)
    DATUV_col = findColumnLetterByColNameAndStartRow(data_ws, "DATUV", DATA_HEADER_ROW)
    MATNR_col = findColumnLetterByColNameAndStartRow(data_ws, "MATNR", DATA_HEADER_ROW)
    WERKS_A_col = findColumnLetterByColNameAndStartRow(data_ws, "WERKS_A", DATA_HEADER_ROW)
    MEINS_col = findColumnLetterByColNameAndStartRow(data_ws, "MEINS", DATA_HEADER_ROW)
    skip_cond_3 = set()
    for i in range(DATA_ROW_COUNT+1, n_of_data + DATA_ROW_COUNT+1):
        PLNNR = data_ws[PLNNR_col + str(i)].value
        PLNAL = data_ws[PLNAL_col + str(i)].value
        MATNR = data_ws[MATNR_col + str(i)].value
        WERKS_A = data_ws[WERKS_A_col + str(i)].value
        d = dict()
        d["PLNNR"] = PLNNR
        d["PLNAL"] = PLNAL
        d["WERKS_A"] = WERKS_A
        d["MATNR"] = MATNR
        #cond_1 = check_duplicate_key(data_ws, DATA_HEADER_ROW, DATA_ROW_COUNT, d)
        match_cond_1 = find_by_keys(data_ws, DATA_HEADER_ROW, DATA_ROW_COUNT, d)
        # print("Cond1", match_cond_1)

        header_ws = dataWb.get_sheet_by_name("01 - Header")
        d = dict()
        d["PLNNR"] = PLNNR
        d["PLNAL"] = PLNAL
        #cond_2 = is_key_exist(header_ws, 2, 2, d)
        match_cond_2 = find_by_keys(header_ws, 2, 2, d)
        # print("Cond2", match_cond_2)



        DATUV = data_ws[DATUV_col + str(i)].value
        data = [PLNNR, PLNAL, DATUV, MATNR]
        if len(match_cond_1) > 1:
            writeHeaderReport(active_ws, "ERROR", data, ValidateError.DUPLICATE_KEY[1], "N="+str(len(match_cond_1)))
        if len(match_cond_2) < 1:
            writeHeaderReport(active_ws, "ERROR", data, ValidateError.UNDEFINED[1].format("Group not mapping with 01-Header"), "N="+str(len(match_cond_2)))

        # New Condition_3
        if len(skip_cond_3) == 0 or i not in skip_cond_3:
            MEINS = data_ws[MEINS_col + str(i)].value
            d = dict()
            d["PLNNR"] = PLNNR
            d["PLNAL"] = PLNAL
            match_cond_3_1 = find_by_keys(data_ws, DATA_HEADER_ROW, DATA_ROW_COUNT, d)
            d["MEINS"] = MEINS
            match_cond_3_2 = find_by_keys(data_ws, DATA_HEADER_ROW, DATA_ROW_COUNT, d)
            if len(match_cond_3_1) != len(match_cond_3_2):
                writeHeaderReport(active_ws, "ERROR", data, ValidateError.UNDEFINED[1].format("Materials with different units are assigned in same group"), "N=" + str(len(match_cond_3_1) - len(match_cond_3_2)))
            skip_cond_3 = match_cond_3_1.union(match_cond_3_2)
        else:
            match_cond_3_1.discard(i)
            match_cond_3_2.discard(i)
            if i in skip_cond_3:
                skip_cond_3.discard(i)

        #if cond_1:
            #writeHeaderReport(active_ws, "ERROR", data, ValidateError.DUPLICATE_KEY[1], "row="+str(i))
        #if not cond_2:
            #writeHeaderReport(active_ws, "ERROR", data, ValidateError.UNDEFINED[1].format("Group does not exist in 02 - Operation"), "row="+str(i))
        

    print("Fin Additional Condition")

    # Check By Field
    key = ["PLNNR", "PLNAL"]
    for i in range(DATA_ROW_COUNT+1, n_of_data + DATA_ROW_COUNT+1):
        for j in range(1, data_ws.max_column +1):
            report_data = [
                data_ws[PLNNR_col+str(i)].value, 
                data_ws[PLNAL_col+str(i)].value,
                data_ws[DATUV_col+str(i)].value,
                data_ws[MATNR_col+str(i)].value
            ]
            key_value = [
                data_ws[PLNNR_col+str(i)].value, 
                data_ws[PLNAL_col+str(i)].value
            ]
            key_data_dict = dict(itertools.izip(key,key_value))
            field_descr = data_ws.cell(row=DATA_FIELD_ROW, column=j).value

            real_data = data_ws.cell(row=i, column=j).value
            if isinstance(real_data, int) or isinstance(real_data, long) or isinstance(real_data, float):
                data = str(real_data)
            else:
                data = real_data
                
            if data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "PLNNR":                
                if isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i)
                if not isNull(data) and len(data) > 15:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "PLNAL":
                if isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i)
                if not isNumeric(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i)
                if not isNull(data) and len(data) > 2:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "WERKS_A":
                if isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i)
                elif not isNull(data) and len(data) > 4:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i)
                elif varB == "Factory" and not isNull(data) and find_in_dict("03-Plant", 1, real_data) is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE_EMPTY[1].format(field_descr), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "DATUV":
                #if isNull(data):
                    #writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i)
                if data != "01012017":
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "01012017"), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "MATNR":
                if isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i)
                if not isNull(data) and len(data) > 18:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "MEINS":
                if isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i)
                if not isNull(data) and len(data) > 6:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i)
            # else:
            #     writeHeaderReport(active_ws, "", report_data, "Success")