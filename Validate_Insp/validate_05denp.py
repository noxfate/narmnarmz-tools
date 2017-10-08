from common import *
from openpyxl.utils import get_column_letter
import openpyxl
from validate import ValidateError, find_in_dict

def writeHeaderReport(ws, status, data, errorMsg, debug=None):
    new_row = []
    new_row.append(status)
    if len(data) != 5:
        raise ValueError("[05-Denp. Char.] Data size is not correct")
    new_row += data
    new_row.append(errorMsg)
    new_row.append(debug)
    insert_new_row(ws, new_row) 

def check_same_matassign_by_MATNR(dataWb, keyDict, data):
    ws = dataWb.get_sheet_by_name("04 - Mat. Assign")
    keys = dict(keyDict)
    keys.pop("VORNR")
    keys.pop("MERKNR")
    keys.pop("ZUORDNR")
    found = find_by_keys(ws, 2, 2, keys)
    found = list(found)
    found.sort()
    row = found[0] if len(found) >= 1 else 0
    if row == 0:
        return False
    col = findColumnLetterByColNameAndStartRow(ws, "MATNR", 2)
    WERKS = ws[col + str(row)].value
    if data == WERKS:
        return True
    return False 

def get_mic_cellvalue_by_colname(dataWb, colname, keyDict):
    MIC_HEADER_ROW = 3
    keys = dict(keyDict)
    keys.pop("ZUORDNR")
    found = find_by_keys(ws, 2, 2, keys)
    found = list(found)
    found.sort()
    row = found[0] if len(found) >= 1 else 0
    if row == 0:
        return None
    mic_ws = dataWb.get_sheet_by_name("03 - MIC")
    col = findColumnLetterByColNameAndStartRow(mic_ws, colname, MIC_HEADER_ROW)
    return mic_ws[col+str(row)].value

def get_decimal_places(ws, row):
    DENP_HEADER_ROW = 2
    col = findColumnLetterByColNameAndStartRow(ws, "DEC_PLACES", DENP_HEADER_ROW)
    dec = ws[col + str(row)].value
    if dec is None:
        return 0
    return dec

def get_value_by_row_colname(ws, colname, row):
    DENP_HEADER_ROW = 2
    col = findColumnLetterByColNameAndStartRow(ws, colname, DENP_HEADER_ROW)
    return ws[col + str(row)].value

def validate(wb, dataWb):
    ## CONFIG HERE NA N'Narm ##
    DATA_TAB_NAME = "05 - Denp. Char." # sheet name to find data
    DATA_ROW_COUNT = 2 # how many row to skip in header
    DATA_FIELD_ROW = 1
    DATA_HEADER_ROW = 2 # what row to find by field
    ROW_START = 2 # row to start writing data
    IS_FREEZE = True # wanna freeze header ?

    active_ws = wb.get_sheet_by_name("05 - Denp. Char.")
    if (IS_FREEZE):
        active_ws.freeze_panes = "A"+ str(ROW_START)
    data_ws = dataWb.get_sheet_by_name(DATA_TAB_NAME)
    n_of_data = data_ws.max_row - DATA_ROW_COUNT

    # CHECK Addtional Condition 1-2
    PLNNR_col = findColumnLetterByColNameAndStartRow(data_ws, "PLNNR", DATA_HEADER_ROW)
    PLNAL_col = findColumnLetterByColNameAndStartRow(data_ws, "PLNAL", DATA_HEADER_ROW)
    VORNR_col = findColumnLetterByColNameAndStartRow(data_ws, "VORNR", DATA_HEADER_ROW)
    MERKNR_col = findColumnLetterByColNameAndStartRow(data_ws, "MERKNR", DATA_HEADER_ROW)
    ZUORDNR_col = findColumnLetterByColNameAndStartRow(data_ws, "ZUORDNR", DATA_HEADER_ROW)
    for i in range(DATA_ROW_COUNT+1, n_of_data + DATA_ROW_COUNT+1):
        PLNNR = data_ws[PLNNR_col + str(i)].value
        PLNAL = data_ws[PLNAL_col + str(i)].value
        VORNR = data_ws[VORNR_col + str(i)].value
        MERKNR = data_ws[MERKNR_col + str(i)].value
        ZUORDNR = data_ws[ZUORDNR_col + str(i)].value
        d = dict()
        d["PLNNR"] = PLNNR
        d["PLNAL"] = PLNAL
        d["VORNR"] = VORNR
        d["MERKNR"] = MERKNR
        d["ZUORDNR"] = ZUORDNR
        match_cond_1 = find_by_keys(data_ws, DATA_HEADER_ROW, DATA_ROW_COUNT, d)
        # print("Cond1", match_cond_1)

        mic_ws = dataWb.get_sheet_by_name("03 - MIC")
        d = dict()
        d["PLNNR"] = PLNNR
        d["PLNAL"] = PLNAL
        d["VORNR"] = VORNR
        d["MERKNR"] = MERKNR
        match_cond_2 = find_by_keys(mic_ws, 3, 3, d)
        # print("Cond2", match_cond_2)

        data = [PLNNR, PLNAL, VORNR, MERKNR, ZUORDNR]
        if len(match_cond_1) > 1:
            writeHeaderReport(active_ws, "ERROR", data, ValidateError.DUPLICATE_KEY[1], "N="+str(len(match_cond_1)))
        if len(match_cond_2) < 1:
            writeHeaderReport(active_ws, "ERROR", data, ValidateError.UNDEFINED[1].format("Group does not exist in 03 - MIC"), "N="+str(len(match_cond_2)))
    
    print("Fin Additional Condition")

    # Check By Field
    key = ["PLNNR", "PLNAL", "VORNR", "MERKNR", "ZUORDNR"]
    for i in range(DATA_ROW_COUNT+1, n_of_data + DATA_ROW_COUNT+1):
        report_data = [
                data_ws[PLNNR_col+str(i)].value, 
                data_ws[PLNAL_col+str(i)].value,
                data_ws[VORNR_col+str(i)].value,
                data_ws[MERKNR_col+str(i)].value,
                data_ws[ZUORDNR_col+str(i)].value
            ]
        key_value = [
                data_ws[PLNNR_col+str(i)].value, 
                data_ws[PLNAL_col+str(i)].value,
                data_ws[VORNR_col+str(i)].value,
                data_ws[MERKNR_col+str(i)].value,
                data_ws[ZUORDNR_col+str(i)].value
            ]
        key_data_dict = dict(itertools.izip(key,key_value))

        
        QPMK_WERKS = get_mic_cellvalue_by_colname(dataWb, "QPMK_WERKS", key_data_dict)
        VERWMERKM = get_mic_cellvalue_by_colname(dataWb, "VERWMERKM", key_data_dict)
        MIC_Checklist = QPMK_WERKS.strip() + VERWMERKM.strip()
        found_QL = find_in_dict("06-MICQL", 1, MIC_Checklist)
        found_QN = find_in_dict("06-MICQN", 1, MIC_Checklist)
        if found_QL is None and found_QN is not None:
            isQL = False
        elif found_QL is not None and found_QN is None:
            isQL = True
        elif found_QL is None and found_QN is None:
            continue
        else:
            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Found in both MICQL, MICQN"), i)
            continue
        
        for j in range(1, data_ws.max_column +1):
            
            field_descr = data_ws.cell(row=DATA_FIELD_ROW, column=j).value

            real_data = data_ws.cell(row=i, column=j).value
            if isinstance(real_data, int) or isinstance(real_data, long) or isinstance(real_data, float):
                data = str(real_data)
            else:
                data = real_data
                
            if data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "PLNNR":                
                if data is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i)
                if data is not None and len(data) > 8:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGHT[1].format(field_descr), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "PLNAL":
                if data is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i)
                if not isNumeric(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i)
                if data is not None and len(data) > 2:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGHT[1].format(field_descr), i)            
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "VORNR":
                if data is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i)
                if data is not None and len(data) > 4:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGHT[1].format(field_descr), i)
                if not isNumOnly(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "MERKNR":
                if data is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i)
                if data is not None and len(data) > 4:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGHT[1].format(field_descr), i)
                if not isNumOnly(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "ZUORDNR":
                if data is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i)
                if data is not None and len(data) > 4:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGHT[1].format(field_descr), i)
                if not isNumeric(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "MATNR":
                if data is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i)
                if data is not None and len(data) > 40:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGHT[1].format(field_descr), i)
                if not check_same_matassign_by_MATNR(dataWb, key_data_dict, real_data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Material doesn't exist in Material Assignment"), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "MATKX":
                if data is not None and len(data) > 40:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGHT[1].format(field_descr), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "WERKS_A":
                if data is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i)
                if data is not None and len(data) > 4:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGHT[1].format(field_descr), i)
                if not find_in_dict("03-Plant", 1, real_data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE_EMPTY[1].format(field_descr), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "LIFNR":
                if data is not None and len(data) > 10:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGHT[1].format(field_descr), i) 
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "DEC_PLACES":
                if isQL:
                    if data is not None:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Qualitative: DEC_PLACES must be blank"), i) 
                else:
                    if data is not None and not isNumeric(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i) 
                    if data is not None and len(data) > 3:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGHT[1].format(field_descr), i)             
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "MEAS_UNIT":
                if isQL:
                    if data is not None:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Qualitative: MEAS_UNIT must be blank"), i) 
                else:
                    if data is not None and len(data) > 6:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGHT[1].format(field_descr), i)            
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "SOLLWERT":
                if isQL:
                    if data is not None:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Qualitative: Target Value must be blank"), i) 
                else:
                    if data is not None and not isNumeric(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i) 
                    if data is not None and len(data) > 16:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGHT[1].format(field_descr), i)
                    dec = get_decimal_places(data_ws, i)
                    if not checkDecimalPlace(dec, real_data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Target Value conflict with Decimal place"), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "TOLERANZUN":
                if isQL:
                    if data is not None:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Qualitative: Lower Limit must be blank"), i) 
                else:
                     if get_mic_cellvalue_by_colname(dataWb, "LW_TOL_LMT_IND", key_data_dict) is not None:
                        if data is None or not isNumeric(data) or len(data) > 16:
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Conflict with LW_TOL_LMT_IND"), i)             
                        dec = get_decimal_places(data_ws, i)
                        if not checkDecimalPlace(dec, real_data):
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Lower Limit conflict with Decimal place"), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "TOLERANZOB":
                if isQL:
                    if data is not None:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Qualitative: Lower Limit must be blank"), i) 
                else:
                    if get_mic_cellvalue_by_colname(dataWb, "UP_TOL_LMT_IND", key_data_dict) is not None:
                        if data is None or not isNumeric(data) or len(data) > 16:
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Conflict with UP_TOL_LMT_IND"), i)             
                        dec = get_decimal_places(data_ws, i)
                        if not checkDecimalPlace(dec, real_data):
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Upper Limit conflict with Decimal place"), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "AUSWMENGE1":
                if isQL:
                    if data is None:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Selected set cannot be blank"), i) 
                    if data is not None and len(data) > 8:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGHT[1].format(field_descr), i) 
                    QWERKAUSW = get_value_by_row_colname(data_ws, "QWERKAUSW", i)
                    x = QWERKAUSW.strip() + data.strip()
                    if find_in_dict("08-Selected Set", 1, x) is None:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE_EMPTY[1].format(field_descr), i) 
                else:
                    if data is not None:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Quantitative: Selected set must be blank"), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "QWERKAUSW":
                if isQL:
                    if data is None:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Plant for Selected set cannot be blank"), i) 
                    if data is not None and len(data) > 4:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGHT[1].format(field_descr), i) 
                    AUSWMENGE1 = get_value_by_row_colname(data_ws, "AUSWMENGE1", i)
                    x = data.strip() + AUSWMENGE1.strip()
                    if find_in_dict("08-Selected Set", 1, x) is None:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE_EMPTY[1].format(field_descr), i) 
                else:
                    if data is not None:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Quantitative: Plant for Selected set must be blank"), i)