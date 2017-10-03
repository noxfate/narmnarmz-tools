from common import *
from openpyxl.utils import get_column_letter
import openpyxl
from validate import ValidateError

def writeHeaderReport(ws, status, data, errorMsg):
    new_row = []
    new_row.append(status)
    if len(data) != 4:
        raise ValueError("[01-Header] Data size is not correct")
    new_row += data
    new_row.append(errorMsg)
    insert_new_row(ws, new_row)
    

def validate(wb, dataWb):
    ## CONFIG HERE NA N'Narm ##
    DATA_TAB_NAME = "01 - Header" # sheet name to find data
    DATA_ROW_COUNT = 2 # how many row to skip in header
    DATA_FIELD_ROW = 1
    DATA_HEADER_ROW = 2 # what row to find by field
    ROW_START = 2 # row to start writing data
    IS_FREEZE = True # wanna freeze header ?

    active_ws = wb.get_sheet_by_name("01 - Header")
    if (IS_FREEZE):
        active_ws.freeze_panes = "A"+ str(ROW_START)
    data_ws = dataWb.get_sheet_by_name(DATA_TAB_NAME)
    n_of_data = data_ws.max_row - DATA_ROW_COUNT

    # CHECK Addtional Condition 1-2
    PLNNR_col = findColumnLetterByColNameAndStartRow(data_ws, "PLNNR", DATA_HEADER_ROW)
    PLNAL_col = findColumnLetterByColNameAndStartRow(data_ws, "PLNAL", DATA_HEADER_ROW)
    WERKS_col = findColumnLetterByColNameAndStartRow(data_ws, "WERKS", DATA_HEADER_ROW)
    KTEXT_col = findColumnLetterByColNameAndStartRow(data_ws, "KTEXT", DATA_HEADER_ROW)
    for i in range(ROW_START, n_of_data + ROW_START):
        PLNNR = data_ws[PLNNR_col + str(i)].value
        found_PLNNR = findCellListInColumnByValue(data_ws, "PLNNR", PLNNR, DATA_HEADER_ROW)
        PLNAL = data_ws[PLNAL_col + str(i)].value
        found_PLNAL = findCellListInColumnByValue(data_ws, "PLNAL", PLNAL, DATA_HEADER_ROW)
        match_cond_1 = set(found_PLNNR) & set(found_PLNAL)

        VERWE_col = findColumnLetterByColNameAndStartRow(data_ws, "VERWE", DATA_HEADER_ROW)
        VERWE = data_ws[VERWE_col+str(i)]
        found_VERWE = findCellListInColumnByValue(data_ws, "VERWE", VERWE, DATA_HEADER_ROW)
        match_cond_2 = set(found_VERWE) & set(found_PLNNR)

        WERKS = data_ws[WERKS_col + str(i)].value
        KTEXT = data_ws[KTEXT_col + str(i)].value
        if len(match_cond_1) > 1 or i not in match_cond_1:
            data = [PLNNR, PLNAL, WERKS, KTEXT]
            writeHeaderReport(active_ws, "ERROR", data, ValidateError.DUPLICATE_KEY[1])
        if len(match_cond_2) > 1 or i not in match_cond_2:
            data = [PLNNR, PLNAL, WERKS, KTEXT]
            writeHeaderReport(active_ws, "ERROR", data, ValidateError.DUPLICATE_KEY[1])
    
    # Check By Field
    for i in range(ROW_START, n_of_data + ROW_START):
        for j in range(1, data_ws.max_column +1):
            report_data = [
                data_ws[PLNNR_col+str(i)], 
                data_ws[PLNAL_col+str(i)],
                data_ws[WERKS_col+str(i)],
                data_ws[KTEXT_col+str(i)]
            ]
            if data_ws.cell(row=DATA_HEADER_ROW, col=j).value == "PLNNR":
                data = data_ws.cell(row=i, col=j).value
                if data is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(data_ws.cell(row=DATA_FIELD_ROW, col=j)))
                if not isChar(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(data_ws.cell(row=DATA_FIELD_ROW, col=j)))
                if len(data) > 8:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGHT[1].format(data_ws.cell(row=DATA_FIELD_ROW, col=j)))

                    