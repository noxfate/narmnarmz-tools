import Tkinter
import tkFileDialog
import os
import openpyxl
import itertools
import easygui
import time
from openpyxl.utils import get_column_letter
from common import *
import sys, traceback

def enum(**enums):
    return type('Enum', (), enums)

ValidateError = enum(
        NOT_NULL=["Not null", "{} cannot be null"],
        NULL=["Null", "Leave {} blank"],
        VALUE_TYPE=["Value type", "{} is incorrect data"],
        LENGTH=["Length", "{} is out of length"],
        FIXED_VALUE=["Fixed value", "{} must be {}"],
        FIXED_VALUE_EMPTY=["Fixed value Not Found in Dict", "{} doesn't exist"],
        FIXED_VALUE_X=["Fixed value X", "X must be capital letter"],
        DUPLICATE_KEY=["Duplicate", "Duplicate code"],
        DUPLICATE=["Duplicate", "Duplicate material assignment"],
        UNDEFINED=["Undefined", "{}"]
    )

cur_path = os.path.dirname(__file__)
new_unit_path = os.path.join(cur_path,'..', 'resource','Dict', 'unit.xlsx')
#new_val_dict_path = os.path.join(cur_path, '..','resource','Dict', '02 Dictionary V1.0.XLSX')
#val_dict_wb = openExcelFile(new_val_dict_path)

def find_in_dict(sheetName, colNumber, input):
    global val_dict_wb
    ws = val_dict_wb.get_sheet_by_name(sheetName)
    if input is not None:
        found = findCellInColumnByValue(ws, colNumber, input, 0)
        # print("Find: ["+sheetName+"] "+str(input), ", Found: "+str(found))
        if found is None:
            return None
        return found
    return None

def find_multiple_in_dict(sheetName, inputDict):
    global val_dict_wb
    ws = val_dict_wb.get_sheet_by_name(sheetName)
    found = []
    if inputDict is not None:
        for k in inputDict.keys():
            cells = findCellListInColumnByValue(ws, k, inputDict[k], 1)
            rows = []
            if cells is not None:
                cells_list = list(cells)
                for i in range(len(cells_list)):
                    rows.append(cells_list[i].row)
            found.append(set(rows))
        result = set.intersection(*found)
        return result
    return set()


import validate_Recipe

def run(varB):

    filePath = openDialog()
    # filePath = "D:/project/narmnarmz-tools/resource/TMP_QM12_Inspection Plan.xlsx"
    start_time = time.time()

    if varB == "Factory":
        new_val_dict_path = os.path.join(cur_path, '..','resource','Dict', '02 Dictionary V1.0.XLSX')
    else:
        new_val_dict_path = os.path.join(cur_path, '..','resource','Dict', '02 Dictionary Farm.xlsx')

    global val_dict_wb
    val_dict_wb = openExcelFile(new_val_dict_path)

    try:
        file_structure = configFileStructure()
        data = openExcelFile(filePath)        
        output_filename = composeFileName(filePath)
        newValidateInspExcel(file_structure, data, output_filename, varB)
        print("--- %s seconds ---" % (time.time() - start_time))
        easygui.msgbox("Your output is "+output_filename+", which is in the same directory that your selected file. \n\nGood luck, have fun!!\n\nExecutime (s): "+str((time.time() - start_time)), title="Success!")
    except TypeError:
        err = traceback.format_exc()
        print(err)
        easygui.msgbox("TypeError: Maybe this happen because the program can't find field in Excel\n\n"+str(err))
    except UnitConversionError as dic:
        easygui.msgbox("Value Not Found: cannot find this following data in 02 Dictionary V1.0.xlsx\n\n"+str(dic))
    except:
        err = traceback.format_exc()
        print(err)
        easygui.msgbox("Unexpected Error: "+str(err))


def composeFileName(fileFullPath):
    return "ERR_"+os.path.basename(fileFullPath)

def newValidateInspExcel(structure, datamodelwb, fileName, varB):
    wb = openpyxl.Workbook()
    old_sheet_list = wb.get_sheet_names()
    for i in structure:
        wb.create_sheet(title=i)
        sheet = wb.get_sheet_by_name(i)
        for j in range(len(structure[i])):
            sheet[get_column_letter(j+1)+'1'] = structure[i][j]
    for i in old_sheet_list:
        wb.remove_sheet(wb.get_sheet_by_name(i))    

    print("....Start Building....")

    print("....Validate Recipe....")
    validate_Recipe.validate(wb, datamodelwb, varB)
        
    print("Output: ", fileName)
    wb.save(fileName)

def configFileStructure():
    output_sheets = ["QM_Recipe"] # index (order) DOES MATTER!!
    main_header = []
    QM_Recipe = ["Status", "Group", "Group Counter", "Operation", "Char. No.", "Error Message"]

    # append order MUST match sheet names in [output_sheets]
    main_header.append(QM_Recipe)
    result = dict(itertools.izip(output_sheets, main_header))
    return result
