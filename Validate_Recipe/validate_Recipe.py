from common import *
from openpyxl.utils import get_column_letter
import openpyxl
from validateR import ValidateError, find_in_dict, find_multiple_in_dict

def writeHeaderReport(ws, status, data, errorMsg, debug=None, colName=None, isQL=None):
    new_row = []
    new_row.append(status)
    if len(data) != 4:
        raise ValueError("[Recipe] Data size is not correct")
    new_row += data
    new_row.append(errorMsg)
    new_row.append(debug)
    new_row.append(colName)
    new_row.append(isQL)
    insert_new_row(ws, new_row) 

def get_decimal_places(ws, row):
    MIC_HEADER = 2
    col = findColumnLetterByColNameAndStartRow(ws, "DEC_PLACES", MIC_HEADER)
    dec = ws[col + str(row)].value
    if dec is None:
        return 0
    return dec

def get_value_by_row_colname(ws, colname, row):
    MIC_HEADER = 2
    col = findColumnLetterByColNameAndStartRow(ws, colname, MIC_HEADER)
    if ws[col+str(row)].value == '':
        return None
    return ws[col + str(row)].value

def validate(wb, dataWb):
    ## CONFIG HERE NA N'Narm ##
    DATA_TAB_NAME = "QM_Recipe" # sheet name to find data
    DATA_ROW_COUNT = 2 # how many row to skip in header
    DATA_FIELD_ROW = 1
    DATA_HEADER_ROW = 2 # what row to find by field
    ROW_START = 2 # row to start writing data
    IS_FREEZE = True # wanna freeze header ?

    active_ws = wb.get_sheet_by_name("QM_Recipe")
    if (IS_FREEZE):
        active_ws.freeze_panes = "A"+ str(ROW_START)
    data_ws = dataWb.get_sheet_by_name(DATA_TAB_NAME)
    n_of_data = data_ws.max_row - DATA_ROW_COUNT
    
    # CHECK Addtional Condition 1-2
    PLNNR_col = findColumnLetterByColNameAndStartRow(data_ws, "PLNNR", DATA_HEADER_ROW)
    PLNAL_col = findColumnLetterByColNameAndStartRow(data_ws, "PLNAL", DATA_HEADER_ROW)
    VORNR_col = findColumnLetterByColNameAndStartRow(data_ws, "VORNR", DATA_HEADER_ROW)
    MERKNR_col = findColumnLetterByColNameAndStartRow(data_ws, "MERKNR", DATA_HEADER_ROW)
    
    for i in range(DATA_ROW_COUNT+1, n_of_data + DATA_ROW_COUNT+1):
        PLNNR = data_ws[PLNNR_col + str(i)].value
        PLNAL = data_ws[PLNAL_col + str(i)].value
        VORNR = data_ws[VORNR_col + str(i)].value
        MERKNR = data_ws[MERKNR_col + str(i)].value
        d = dict()
        d["PLNNR"] = PLNNR
        d["PLNAL"] = PLNAL
        d["VORNR"] = VORNR
        d["MERKNR"] = MERKNR
        match_cond_1 = find_by_keys(data_ws, DATA_HEADER_ROW, DATA_ROW_COUNT, d)
        # print("Cond1", match_cond_1)

        data = [PLNNR, PLNAL, VORNR, MERKNR]
        if len(match_cond_1) > 1:
            writeHeaderReport(active_ws, "ERROR", data, ValidateError.DUPLICATE_KEY[1], "N="+str(len(match_cond_1)))
    
    print("Fin Additional Condition")
	
    # Check By Field
    key = ["PLNNR", "PLNAL", "VORNR", "MERKNR"]
    for i in range(DATA_ROW_COUNT+1, n_of_data + DATA_ROW_COUNT+1):
        report_data = [
                data_ws[PLNNR_col+str(i)].value, 
                data_ws[PLNAL_col+str(i)].value,
                data_ws[VORNR_col+str(i)].value,
                data_ws[MERKNR_col+str(i)].value,
            ]
        key_value = [
                data_ws[PLNNR_col+str(i)].value, 
                data_ws[PLNAL_col+str(i)].value,
                data_ws[VORNR_col+str(i)].value,
                data_ws[MERKNR_col+str(i)].value,
            ]
        key_data_dict = dict(itertools.izip(key,key_value))

        QPMK_WERKS_col = findColumnLetterByColNameAndStartRow(data_ws, "QPMK_WERKS", DATA_HEADER_ROW)
        VERKMERKM_col = findColumnLetterByColNameAndStartRow(data_ws, "VERWMERKM", DATA_HEADER_ROW)
        QPMK_WERKS = data_ws[QPMK_WERKS_col + str(i)].value
        VERWMERKM = data_ws[VERKMERKM_col + str(i)].value
        if QPMK_WERKS is None:
            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format("Plant - MIC"), i)
        if VERWMERKM is None:
            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format("Master Inspection Characteristic"), i)

        x_dict = dict()
        x_dict[2] = QPMK_WERKS
        x_dict[3] = VERWMERKM
        found_QL = find_multiple_in_dict("06-MICQL", x_dict)
        found_QN = find_multiple_in_dict("06-MICQN", x_dict)
        if len(found_QL) == 0 and len(found_QN) != 0:
            isQL = False
        elif len(found_QL) != 0 and len(found_QN) == 0:
            isQL = True
        elif len(found_QL) == 0 and len(found_QN) == 0:
            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Doesn't exist"), i)
            continue
        #else:
            #writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Found in both MICQL, MICQN"), i)
            #continue

        for j in range(1, data_ws.max_column +1):
            field_descr = data_ws.cell(row=DATA_FIELD_ROW, column=j).value
            real_data = data_ws.cell(row=i, column=j).value
            if isinstance(real_data, int) or isinstance(real_data, long) or isinstance(real_data, float):
                data = str(real_data)
            else:
                data = real_data
            
            if data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "PLNNR": #Group           
                if isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                elif not isNull(data) and len(data) > 15:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "PLNAL": #Group Counter
                #if isNull(data):
                    #writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                if not isNull(data) and not isNumeric(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                if not isNull(data) and len(data) > 2:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "WERKS": #Plant
            	if isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                elif not isNull(data) and len(data) > 4:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                elif not isNull(data) and find_in_dict("03-Plant",1, real_data) is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE_EMPTY[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)            
            #elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "DATUV": #Valid date
                #if isNull(data):
                    #writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                #elif not isNull(data) and data != "01012017":
                    #writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "01012017"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "SLWBEZ": #Inspection Point
                if isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                elif not isNull(data) and data != "FH1":
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "FH1"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "PARL": #Partial-lot assign.
                if not isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "QPRZIEHVER": #Sample-drawing proc.
                if not isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "DMYM": #Dynamic mod. level
                if not isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "QDYNREGEL": #Modification rule
                if not isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "VORNR":
                if isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                elif not isNull(data) and len(data) != 4:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Operation should have 4 digits"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                elif not isNumOnly(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                elif int(data) < 1010:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Operation of Recipe start at 1010"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "STEUS": #Status
                if isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                elif not isNull(data) and data != "QM02":
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "QM02"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "LTXA1": #Status
                if isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                if not isNull(data) and len(data) > 40:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "MERKNR":
                if isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                elif not isNull(data) and len(data) != 4:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Characteristic number should have 4 digits"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                elif not isNumOnly(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "MKVERSION":
                if not isNull(data) and data != "1":
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "1"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "KURZTEXT":
                if not isNull(data) and len(data) > 40:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "PMETHODE":
                if not isNull(data) and len(data) > 8:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                if not isNull(data) and find_in_dict("09-Method",3, real_data) is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE_EMPTY[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "QMTB_WERKS":
                if not isNull(data) and get_value_by_row_colname(data_ws, "PMETHODE", i) is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("PMETHODE conflict with QMTB_WERKS"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                if not isNull(data) and len(data) > 4:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                if not isNull(data) and find_in_dict("09-Method",2, real_data) is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE_EMPTY[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "PMTVERSION":
                if not isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "TARGET_VAL_CHECK":
                if isQL:
                    if not isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Qualitative: TARGET_VAL_CHECK must be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                else:
                    if get_value_by_row_colname(data_ws, "SOLLWERT", i) is not None:
                        if isNull(data):
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                        if not isNull(data) and data != "X":                    
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Conflict with Target Value"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "LW_TOL_LMT_IND":
                if isQL:
                    if not isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Qualitative: LW_TOL_LMT_IND must be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                else:
                    if get_value_by_row_colname(data_ws, "TOLERANZUN", i) is not None:
                        if isNull(data):
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                        if not isNull(data) and data != "X":                    
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Conflict with Lower Limit"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "UP_TOL_LMT_IND":
                if isQL:
                    if not isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Qualitative: UP_TOL_LMT_IND must be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                else:
                    if get_value_by_row_colname(data_ws, "TOLERANZOB", i) is not None:
                        if isNull(data):
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                        if not isNull(data) and data != "X":                    
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Conflict with Upper Limit"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "DEC_PLACES":
                if isQL:
                    if not isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Qualitative: DEC_PLACES must be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                else:
                    if not isNull(data) and not isNumeric(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    if not isNull(data) and len(data) > 3:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "SOLLWERT":
                if isQL:
                    if not isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Qualitative: Target Value must be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                else:
                    if not isNull(data) and not isNumeric(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    elif not isNull(data) and len(data) > 16:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    elif get_value_by_row_colname(data_ws, "TARGET_VAL_CHECK", i) is not None:
                        if isNull(data):
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Conflict with TARGET_VAL_CHECK"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                        dec = get_decimal_places(data_ws, i)
                        if not checkDecimalPlace(dec, data):
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Target Value conflict with Decimal place"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                        if not isNumeric(data):
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                        v1 = get_value_by_row_colname(data_ws, "TOLERANZUN", i) # lower
                        v2 = get_value_by_row_colname(data_ws, "TOLERANZOB", i) # upper
                        if not isNull(real_data):
                            try:
                                if v1 is not None:
                                    if float(real_data) < float(v1):
                                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Target Value conflict with Lower Limit"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                                if v2 is not None:
                                    if float(real_data) > float(v2):
                                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Target Value conflict with Upper Limit"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                            except ValueError:
                                writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format("SOLLWERT, TOLERANZUN, TOLERANZOB"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)

            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "TOLERANZUN":
                if isQL:
                    if not isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Qualitative: Lower Limit must be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                else:
                    if not isNull(data) and not isNumeric(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    elif not isNull(data) and len(data) > 16:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    elif get_value_by_row_colname(data_ws, "LW_TOL_LMT_IND", i) is not None:
                        if isNull(data):
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Conflict with LW_TOL_LMT_IND"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                        if not isNumeric(data):
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                        dec = get_decimal_places(data_ws, i)
                        if not checkDecimalPlace(dec, data):
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Lower Limit conflict with Decimal place"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                        v1 = get_value_by_row_colname(data_ws, "SOLLWERT", i) # target
                        v2 = get_value_by_row_colname(data_ws, "TOLERANZOB", i) # upper
                        if not isNull(real_data):    
                            try:
                                if v1 is not None:
                                    if float(real_data )> float(v1):
                                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Lower Limit conflict with Target Value"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                                if v2 is not None:
                                    if float(real_data )> float(v2):
                                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Lower Limit conflict with Upper Value"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                            except ValueError:
                                writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format("SOLLWERT, TOLERANZUN, TOLERANZOB"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "TOLERANZOB":
                if isQL:
                    if not isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Qualitative: Upper Limit must be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                else:
                    if not isNull(data) and not isNumeric(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    elif not isNull(data) and len(data) > 16:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    elif get_value_by_row_colname(data_ws, "UP_TOL_LMT_IND", i) is not None:
                        if isNull(data):
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Conflict with UP_TOL_LMT_IND"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                        if not isNumeric(data):
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                        dec = get_decimal_places(data_ws, i)
                        if not checkDecimalPlace(dec, data):
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Upper Limit conflict with Decimal place"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                        v1 = get_value_by_row_colname(data_ws, "SOLLWERT", i) # target
                        v2 = get_value_by_row_colname(data_ws, "TOLERANZUN", i) # lower
                        if not isNull(real_data):
                            try:
                                if v1 is not None:
                                    if float(real_data )< float(v1):
                                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Upper Limit conflict with Target Value"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                                if v2 is not None:
                                    if float(real_data )< float(v2):
                                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Upper Limit conflict with Lower Limit"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                            except ValueError:
                                writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format("SOLLWERT, TOLERANZUN, TOLERANZOB"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "SAMPLING_PROCEDU": #Valid date
                if isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                elif not isNull(data) and data != "X":
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "X"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "MEAS_UNIT":
                if isQL:
                    if not isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Qualitative: MEAS_UNIT must be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                else:
                    if not isNull(data) and len(data) > 6:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    if not isNull(data) and find_in_dict("02-Unit", 4, real_data) is None:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MEAS_UNIT doesn't exist"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "SPC_IND":
                if isQL:
                    if not isNull(data):
                         writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Qualitative: SPC Characteristic must be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                else:                    
                    SPC_CRIT = get_value_by_row_colname(data_ws, "SPC_CRITERION_KEY", i)
                    if isNull(SPC_CRIT) and not isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("SPC Characteristic Conflict with SPC criterion"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    elif SPC_CRIT is not None and data != "X":
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "X"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "SPC_CRITERION_KEY":
                if isQL:
                    if not isNull(data):
                         writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Qualitative: SPC criterion must be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                else:
                    SPC_IND = get_value_by_row_colname(data_ws, "SPC_IND", i)
                    if isNull(SPC_IND) and not isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("SPC Characteristic Conflict with SPC criterion"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    elif SPC_IND is not None and data != "070":
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "070"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "RESULT_RECORDING_SUM":
            	RESULT_RECORDING_SIN = get_value_by_row_colname(data_ws, "RESULT_RECORDING_SIN", i)
            	if not isNull(data) and data != "X":                    
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "X"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                if isNull(data) and isNull(RESULT_RECORDING_SIN):
                	writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Mark 'X' either Summ. Recording or Single Result"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                if not isNull(data) and not isNull(RESULT_RECORDING_SIN):
                	writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Mark 'X' either Summ. Recording or Single Result"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "RESULT_RECORDING_SIN":
            	if not isNull(data) and data != "X":                    
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "X"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "CONFIRMATION_CAT_REQ":
            	CONFIRMATION_CAT_OPT = get_value_by_row_colname(data_ws, "CONFIRMATION_CAT_OPT", i)
            	if not isNull(data) and data != "X":                    
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "X"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                if isNull(data) and isNull(CONFIRMATION_CAT_OPT):
                	writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Mark 'X' either Optional Char. or Required Char."), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                if not isNull(data) and not isNull(CONFIRMATION_CAT_OPT):
                	writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Mark 'X' either Optional Char. or Required Char."), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "CONFIRMATION_CAT_OPT":
            	if not isNull(data) and data != "X":                    
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "X"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "SCOPE_IND_NOF":
            	SCOPE_IND_FIX = get_value_by_row_colname(data_ws, "SCOPE_IND_FIX", i)
            	if not isNull(data) and data != "X":                    
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "X"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                if isNull(data) and isNull(SCOPE_IND_FIX):
                	writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Mark 'X' either Fixed Scope or Scope Not Fixed"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                if not isNull(data) and not isNull(SCOPE_IND_FIX):
                	writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Mark 'X' either Fixed Scope or Scope Not Fixed"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "SCOPE_IND_FIX":
            	SPC_IND = get_value_by_row_colname(data_ws, "SPC_IND", i)
            	if not isNull(data) and data != "X":                    
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "X"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                if isNull(data) and not isNull(SPC_IND):
                	writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("SPC must Fixed Scope"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "LONG_TERM_INSP_I":
            	if not isNull(data) and data != "X":                    
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "X"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "DOCU_REQU":
            	if not isNull(data) and data != "X":                    
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "X"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "MEAS_VALUE_CONFI":
                if isQL:
                    if not isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Qualitative: Record measured vals must be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                else:
                    if isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    if not isNull(data) and data != "X":
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "X"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "DOCU_REQU":
            	if not isNull(data) and data != "X":                    
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "X"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "FORMULA_IND_NO":
            	FORMULA_IND_FO = get_value_by_row_colname(data_ws, "FORMULA_IND_FO", i)
            	if not isNull(data) and data != "X":
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "X"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            	if isQL :
            		if isNull(data):
            			writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Qualitative: Mark 'X' at No Formula"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            	elif not isQL:
            		if isNull(data) == isNull(FORMULA_IND_FO):
            			writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Mark 'X' either No Formula or Calc. charac."), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "FORMULA_IND_FO":
            	FORMULA_FIELD_1 = get_value_by_row_colname(data_ws, "FORMULA_FIELD_1", i)
            	if isQL:
            		if not isNull(data):
            			writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Qualitative: Cannot use Calc. charac."), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            	elif not isQL:
	            	if not isNull(data) and data != "X":
	                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "X"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "FORMULA_FIELD_1":
            	FORMULA_IND_FO = get_value_by_row_colname(data_ws, "FORMULA_IND_FO", i)
            	if not isNull(data) and len(data) > 40:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                if isQL:
                	if not isNull(data):
            			writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Qualitative: Cannot use Formula"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            	elif not isQL:
            		if isNull(data) != isNull(FORMULA_IND_FO) :
            			writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Calc. charac conflict with Formula"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "STICHPRVER":
                MIC_SPC = get_value_by_row_colname(data_ws, "SPC_IND", i)
                if isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                if not isNull(data) and len(data) > 8:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                elif isNull(MIC_SPC):
                    found = find_in_dict("R7-SampPoint", 1, real_data)
                elif not isNull(MIC_SPC):
                    found = find_in_dict("R7-SampPointSPC", 1, real_data)
                if isNull(found):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Incorrect sampling procedure"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "PROBEMGEH":
                if isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                if not isNull(data) and len(data) > 6:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            	if not isNull(data) and find_in_dict("02-Unit", 4, real_data) is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MEAS_UNIT doesn't exist"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "PRUEFEINH":
                if isNull(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                if not isNull(data) and not isNumeric(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                if not isNull(data) and data != "1":
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "1"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "ATTRIBUTE_REQUIR":
                if not isQL:
                    if not isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Quantitative: ATTRIBUTE_REQUIR must be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                else:
                    if isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    if not isNull(data) and data != "X":
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "X"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "AUSWMENGE1":
                if isQL:
                    if isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Qualitative: Selected set cannot be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    if not isNull(data) and len(data) > 8:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    QWERKAUSW = get_value_by_row_colname(data_ws, "QWERKAUSW", i)
                    x_dict = dict()
                    x_dict[3] = real_data
                    x_dict[2] = QWERKAUSW
                    if find_multiple_in_dict("08-Selected Set", x_dict) is None:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE_EMPTY[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                else:
                    if not isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Quantitative: Selected set must be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "QWERKAUSW":
                if isQL:
                    if isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Qualitative: Plant for Selected set cannot be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    if not isNull(data) and len(data) > 4:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                    AUSWMENGE1 = get_value_by_row_colname(data_ws, "AUSWMENGE1", i)
                    x_dict = dict()
                    x_dict[3] = AUSWMENGE1
                    x_dict[2] = real_data
                    if find_multiple_in_dict("08-Selected Set", x_dict) is None:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE_EMPTY[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
                else:
                    if not isNull(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("MIC Quantitative: Plant for Selected set must be blank"), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "DUMMY10":
                if not isNull(data) and len(data) > 10:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "DUMMY20":
                if not isNull(data) and len(data) > 20:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "DUMMY40":
                if not isNull(data) and len(data) > 40:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i, data_ws.cell(row=DATA_HEADER_ROW, column=j).value, isQL)