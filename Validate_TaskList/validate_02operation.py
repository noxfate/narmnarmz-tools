from common import *
from openpyxl.utils import get_column_letter
import openpyxl
from validateTL import ValidateError, find_in_dict, find_multiple_in_dict

def writeHeaderReport(ws, status, data, errorMsg, debug=None):
    new_row = []
    new_row.append(status)
    if len(data) != 4:
        raise ValueError("[02-Operation] Data size is not correct")
    new_row += data
    new_row.append(errorMsg)
    new_row.append(debug)
    insert_new_row(ws, new_row)

def get_value_by_row_colname(ws, colname, row):
    MIC_HEADER = 7
    col = findColumnLetterByColNameAndStartRow(ws, colname, MIC_HEADER)
    return ws[col + str(row)].value 

def validate(wb, dataWb):
    ## CONFIG HERE NA N'Narm ##
    DATA_TAB_NAME = "TaskList Operation" # sheet name to find data
    DATA_ROW_COUNT = 7 # how many row to skip in header
    DATA_FIELD_ROW = 1
    DATA_HEADER_ROW = 7 # what row to find by field
    ROW_START = 2 # row to start writing data
    IS_FREEZE = True # wanna freeze header ?

    active_ws = wb.get_sheet_by_name("2. Task List Operation")
    if (IS_FREEZE):
        active_ws.freeze_panes = "A"+ str(ROW_START)
    data_ws = dataWb.get_sheet_by_name(DATA_TAB_NAME)
    n_of_data = data_ws.max_row - DATA_ROW_COUNT

    # CHECK Addtional Condition 1-2
    PLNNR_col = findColumnLetterByColNameAndStartRow(data_ws, "PLNNR", DATA_HEADER_ROW)
    PLNAL_col = findColumnLetterByColNameAndStartRow(data_ws, "PLNAL", DATA_HEADER_ROW)
    VORNR_col = findColumnLetterByColNameAndStartRow(data_ws, "VORNR", DATA_HEADER_ROW)
    LTXA1_col = findColumnLetterByColNameAndStartRow(data_ws, "LTXA1", DATA_HEADER_ROW)
    STEUS_col = findColumnLetterByColNameAndStartRow(data_ws, "STEUS", DATA_HEADER_ROW)
    for i in range(DATA_ROW_COUNT+1, n_of_data + DATA_ROW_COUNT+1):
        PLNNR = data_ws[PLNNR_col + str(i)].value
        PLNAL = data_ws[PLNAL_col + str(i)].value
        VORNR = data_ws[VORNR_col + str(i)].value
        STEUS = get_value_by_row_colname(data_ws,"STEUS",i)
        #Check Duplicate
        d = dict()
        d["PLNNR"] = PLNNR
        d["PLNAL"] = PLNAL
        d["VORNR"] = VORNR
        match_cond_1 = find_by_keys(data_ws, DATA_HEADER_ROW, DATA_ROW_COUNT, d)
        # print("Cond1", match_cond_1)

        #Check Header
        header_ws = dataWb.get_sheet_by_name("Task List Header")
        d = dict()
        d["PLNNR"] = PLNNR
        d["PLNAL"] = PLNAL
        match_cond_2 = find_by_keys(header_ws, 7, 8, d)
        # print("Cond2", match_cond_2)

        match_cond_3 = ''
        if STEUS == "PM02":
            #Check Planned
            header_ws = dataWb.get_sheet_by_name("Task List Planned Servie")
            d = dict()
            d["PLNNR"] = PLNNR
            d["PLNAL"] = PLNAL
            d["VORNR"] = VORNR
            match_cond_3 = find_by_keys(header_ws, 6, 7, d)
            # print("Cond3", match_cond_3)

            if len(match_cond_3) < 1:
                #check unplanned
                header_ws = dataWb.get_sheet_by_name("Task List Unplanned Service")
                match_cond_3 = find_by_keys(header_ws, 7, 8, d)

        LTXA1 = data_ws[LTXA1_col + str(i)].value
        if len(match_cond_1) > 1:
            data = [PLNNR, PLNAL, VORNR, LTXA1]
            writeHeaderReport(active_ws, "ERROR", data, ValidateError.DUPLICATE_KEY[1], "N="+str(len(match_cond_1)))
        if len(match_cond_2) < 1:
            data = [PLNNR, PLNAL, VORNR, LTXA1]
            writeHeaderReport(active_ws, "ERROR", data, ValidateError.UNDEFINED[1].format("Group not mapping with 1. Task List Header"), "N="+str(len(match_cond_2)))
        if STEUS == "PM02" and len(match_cond_3) < 1:
            data = [PLNNR, PLNAL, VORNR, LTXA1]
            writeHeaderReport(active_ws, "ERROR", data, ValidateError.UNDEFINED[1].format("PM02 must be created in Planned or Unplanned Service"), "N="+str(len(match_cond_2)))

    print("Fin Additional Condition")

    # Check By Field
    key = ["PLNNR", "PLNAL", "VORNR"]
    for i in range(DATA_ROW_COUNT+1, n_of_data + DATA_ROW_COUNT+1):
        for j in range(1, data_ws.max_column +1):
            report_data = [
                data_ws[PLNNR_col+str(i)].value, 
                data_ws[PLNAL_col+str(i)].value,
                data_ws[VORNR_col+str(i)].value,
                data_ws[LTXA1_col+str(i)].value
            ]
            key_value = [
                data_ws[PLNNR_col+str(i)].value, 
                data_ws[PLNAL_col+str(i)].value,
                data_ws[VORNR_col+str(i)].value
            ]
            key_data_dict = dict(itertools.izip(key,key_value))
            field_descr = data_ws.cell(row=DATA_FIELD_ROW, column=j).value

            real_data = data_ws.cell(row=i, column=j).value
            if isinstance(real_data, int) or isinstance(real_data, long) or isinstance(real_data, float):
                data = str(real_data)
            else:
                data = real_data
            
            STEUS = get_value_by_row_colname(data_ws,"STEUS",i) #Control Key
            PLNNR = get_value_by_row_colname(data_ws, "PLNNR", i) #Group

            if data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "PLNNR":                
                if data is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i)
                elif data is not None and len(data) > 15:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "PLNAL":
                if data is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i)
                elif not isNumeric(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i)
                elif data is not None and len(data) > 2:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "VORNR":
                if data is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i)
                if data is not None and len(data) > 4:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i)
                if not isNumOnly(real_data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "LTXA1":
                if data is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i)
                if data is not None and len(data) > 40:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "ARBID":
                if data is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i)
                if data is not None and len(data) > 8:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i)
                if data is not None and find_in_dict("04-Work Center", 3, real_data.upper()) is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE_EMPTY[1].format(field_descr), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "WERKS":
                if data is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i)
                if data is not None and len(data) > 4:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i)
                if data is not None and find_in_dict("04-Work Center", 2, real_data) is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE_EMPTY[1].format(field_descr), i)
                #if not check_same_header_by_werks(dataWb, key_data_dict, real_data):
                    #writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Plants not mapping with Header"), i)            
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "STEUS":
                if data is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i)
                if data != "PM02" and data != "PM01" :
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "PM01 or PM02"), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "ANZZL": #Number of Capacity
                if STEUS == "PM01":
                    if data is None:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i)
                    if data is isNumeric(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i)
                    #if not checkDecimalPlace(1,data):
                        #writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Number of Capacity must be 1 decimal place"), i)
                elif STEUS == "PM02":
                    if data is not None:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NULL[1].format(field_descr), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "ARBEI": #Work
                if STEUS == "PM01":
                    if data is None:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i)
                    if data is isNumeric(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i)
                    #if not checkDecimalPlace(1,data):
                        #writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Work must be 1 decimal place"), i)
                elif STEUS == "PM02":
                    if data is not None:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NULL[1].format(field_descr), i)            
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "ARBEH": #Work Unit
                if STEUS == "PM01":
                    if data != "H":
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "H"), i)
                elif STEUS == "PM02":
                    if data is not None:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("PM02: Work Unit must be blank"), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "DAUNE": # Duration Unit
                if STEUS == "PM01":
                    if data != "H":
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "H"), i)
                elif STEUS == "PM02":
                    if data is not None:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("PM02: Duration Unit must be blank"), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "DAUNE1": #Calculation Key
                if STEUS == "PM01":
                    if data != "2":
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "2"), i)
                elif STEUS == "PM02":
                    if data is not None:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("PM02: Calculation Key must be blank"), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "DAUNO": #Duration
                if STEUS == "PM01":
                    if data is None:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i)
                    if data is isNumeric(data):
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i)
                    #if not checkDecimalPlace(1,data):
                        #writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Duration be 1 decimal place"), i)
                    else:
                        #print(type(str(get_value_by_row_colname(data_ws,"ANZZL",i))))

                        ANZZL = get_value_by_row_colname(data_ws,"ANZZL",i)
                        ARBEI = get_value_by_row_colname(data_ws,"ARBEI",i)

                        if ANZZL is not None and ARBEI is not None:
                            ANZZL = round(float(ANZZL),1)
                            ARBEI = round(float(ARBEI),1)
                            data = round(float(data),1)

                            if ANZZL == 0 and ARBEI == 0 and data ==0:
                                continue
                            elif round(ARBEI/ANZZL,1) != round(data,1):
                                writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Incorrect Duration: Duration must be Work/Number of Capacitie"), round(ARBEI/ANZZL,1))

                elif STEUS == "PM02":
                    if data is not None:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NULL[1].format(field_descr), i)
            
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "LARNT": #Activity type
                if STEUS == "PM01" and str(STEUS)[4:6] == "EC":
                    if data is None:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i)
                    elif find_in_dict("10-Activity Type", 3, data) is None:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Activity Type Doesn't exist"), i)
                elif STEUS == "PM02":
                    if data is not None:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NULL[1].format(field_descr), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "KTSCH": #Standard text key
                if data is not None and data != "CA_REP":
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Standard Text Key must be CA_REP"), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "BMVRG": #Operation Quantity
                if STEUS == "PM01" and data is not None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NULL[1].format(field_descr), i)
                elif STEUS == "PM02" and data != "1":
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "1"), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "BMEIH": #Unit of Measure
                if STEUS == "PM01" and data is not None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NULL[1].format(field_descr), i)
                elif STEUS == "PM02" and data != "AU":
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "AU"), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "PEINH": #Price Unit
                if STEUS == "PM01" and data is not None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NULL[1].format(field_descr), i)
                elif STEUS == "PM02" and data != "1":
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "1"), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "PREIS": #Net Price
                if data is not None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NULL[1].format(field_descr), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "WAERS": #Currency
                if STEUS == "PM01" and data is not None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NULL[1].format(field_descr), i)
                elif STEUS == "PM02" and data != "THB":
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "THB"), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "SAKTO": #Cost Element
                if STEUS == "PM01":
                    if data is not None:
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NULL[1].format(field_descr), i)
                elif STEUS == "PM02":
                    if str(PLNNR)[5] == "C" and data != "5207030020":
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Cost Element Conflit with Group structure"), i)
                    elif str(PLNNR)[5] == "V" and data != "5205020010":
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Cost Element Conflit with Group structure"), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "MATKL": #Material Group
                if STEUS == "PM01" and data is not None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NULL[1].format(field_descr), i)
                elif STEUS == "PM02" and data != "R5006":
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "R5006"), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "EKGRP": #Purchasing Group
                if STEUS == "PM01" and data is not None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NULL[1].format(field_descr), i)
                if STEUS == "PM02":
                    WERKS = get_value_by_row_colname(data_ws,"WERKS",i)
                    x_dict = dict()
                    x_dict[3] = real_data #Purchasing group
                    x_dict[2] = WERKS
                    PurGroup = find_multiple_in_dict("11-Pur Group", x_dict)
                    if PurGroup is None or len(PurGroup) == 0:
                        if find_in_dict("11-Pur Group", 1, real_data) is None:
                            if str(PLNNR)[4:6] == "EC":
                                if  data != "011":
                                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Purchasing Group Conflit with Group structure"), i)
                            elif str(PLNNR)[4:6] != "EC":
                                if  data != "013":
                                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Purchasing Group Conflit with Group structure"), i)
                        else:
                            writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Purchasing Group Conflit with Group structure"), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "LIFNR": #Vendor
                if data is not None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NULL[1].format(field_descr), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "EKORG": #Purchasing org
                if STEUS == "PM01" and data is not None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NULL[1].format(field_descr), i)
                elif STEUS == "PM02" and data != "1000":
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "1000"), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "EBELN": #Agreement
                if data is not None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NULL[1].format(field_descr), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "EBELP": #Item of number
                if data is not None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NULL[1].format(field_descr), i)