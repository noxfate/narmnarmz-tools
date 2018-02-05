from common import *
from openpyxl.utils import get_column_letter
import openpyxl
from validateTL import ValidateError, find_in_dict

def get_value_by_row_colname(ws, colname, row):
    MIC_HEADER = 7
    col = findColumnLetterByColNameAndStartRow(ws, colname, MIC_HEADER)
    return ws[col + str(row)].value

def writeHeaderReport(ws, status, data, errorMsg, debug=None):
    new_row = []
    new_row.append(status)
    if len(data) != 4:
        raise ValueError("[4.2 Task List Unplanned Service] Data size is not correct")
    new_row += data
    new_row.append(errorMsg)
    new_row.append(debug)
    insert_new_row(ws, new_row)  

def get_02Operation_STEUS_by_key(dataWb, keyDict):
    ws = dataWb.get_sheet_by_name("TaskList Operation")
    keys = dict(keyDict)
    #keys.pop("VORNR")
    #keys.pop("MERKNR")
    found = find_by_keys(ws, 7, 8, keys)
    found = list(found)
    found.sort()
    row = found[0] if len(found) >= 1 else 0
    if row == 0:
        return False
    col = findColumnLetterByColNameAndStartRow(ws, "STEUS", 7)
    STEUS = ws[col+str(row)].value
    return STEUS

def validate(wb, dataWb):
    ## CONFIG HERE NA N'Narm ##
    DATA_TAB_NAME = "Task List Unplanned Service" # sheet name to find data
    DATA_ROW_COUNT = 7 # how many row to skip in header
    DATA_FIELD_ROW = 1
    DATA_HEADER_ROW = 7 # what row to find by field
    ROW_START = 2 # row to start writing data
    IS_FREEZE = True # wanna freeze header ?

    active_ws = wb.get_sheet_by_name("4.2 Task List Unplanned Service")
    if (IS_FREEZE):
        active_ws.freeze_panes = "A"+ str(ROW_START)
    data_ws = dataWb.get_sheet_by_name(DATA_TAB_NAME)
    n_of_data = data_ws.max_row - DATA_ROW_COUNT
    
    # CHECK Addtional Condition 1-2
    PLNNR_col = findColumnLetterByColNameAndStartRow(data_ws, "PLNNR", DATA_HEADER_ROW)
    PLNAL_col = findColumnLetterByColNameAndStartRow(data_ws, "PLNAL", DATA_HEADER_ROW)
    VORNR_col = findColumnLetterByColNameAndStartRow(data_ws, "VORNR", DATA_HEADER_ROW)
    SUMLI_col = findColumnLetterByColNameAndStartRow(data_ws, "SUMLIMIT", DATA_HEADER_ROW)
    
    for i in range(DATA_ROW_COUNT+1, n_of_data + DATA_ROW_COUNT+1):
        PLNNR = data_ws[PLNNR_col + str(i)].value
        PLNAL = data_ws[PLNAL_col + str(i)].value
        VORNR = data_ws[VORNR_col + str(i)].value
        d = dict()
        d["PLNNR"] = PLNNR
        d["PLNAL"] = PLNAL
        d["VORNR"] = VORNR

        match_cond_1 = find_by_keys(data_ws, DATA_HEADER_ROW, DATA_ROW_COUNT, d)
        # print("Cond1", match_cond_1)
        
        planned_ws = dataWb.get_sheet_by_name("Task List Planned Servie")
        match_cond_2 = find_by_keys(planned_ws, 6, 7, d)
        #print("Cond2", match_cond_2, len(match_cond_2))

        operation_ws = dataWb.get_sheet_by_name("TaskList Operation")
        match_cond_3 = find_by_keys(operation_ws, 7, 8, d)
        #print("Cond2", match_cond_2, len(match_cond_2))

        SUMLI = data_ws[SUMLI_col + str(i)].value
        data = [PLNNR, PLNAL, VORNR, SUMLI]
        if len(match_cond_1) > 1:           
            writeHeaderReport(active_ws, "ERROR", data, ValidateError.DUPLICATE_KEY[1], "N="+str(len(match_cond_1)))        
        if len(match_cond_2) > 0:
            writeHeaderReport(active_ws, "ERROR", data, ValidateError.UNDEFINED[1].format("Service shouldn't exist in both Planned and Unplanned"), "N="+str(len(match_cond_2)))
        if len(match_cond_3) < 1:
            writeHeaderReport(active_ws, "ERROR", data, ValidateError.UNDEFINED[1].format("Group not mapping with 2. Task List Operation"), "N="+str(len(match_cond_2)))

    print("Fin Additional Condition")

    # Check By Field
    key = ["PLNNR", "PLNAL", "VORNR"]
    #i = row / J= Column
    for i in range(DATA_ROW_COUNT+1, n_of_data + DATA_ROW_COUNT+1):
        for j in range(1, data_ws.max_column +1):
            report_data = [
                data_ws[PLNNR_col+str(i)].value, 
                data_ws[PLNAL_col+str(i)].value,
                data_ws[VORNR_col+str(i)].value,
                data_ws[SUMLI_col+str(i)].value
            ]
            key_value = [
                data_ws[PLNNR_col+str(i)].value, 
                data_ws[PLNAL_col+str(i)].value,
                data_ws[VORNR_col+str(i)].value
            ]
            key_data_dict = dict(itertools.izip(key,key_value))
            field_descr = data_ws.cell(row=DATA_FIELD_ROW, column=j).value

            real_data = data_ws.cell(row=i, column=j).value
            if isinstance(real_data, int) or isinstance(real_data, long) or isinstance(real_data, float):
                data = str(real_data)
            else:
                data = real_data

            STEUS = get_02Operation_STEUS_by_key(dataWb, key_data_dict) #Controy key
            if STEUS == "PM01":
                writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("PM01 mustn't has Service Unplanned"), i)

            if data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "PLNNR":                
                if data is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i)
                if data is not None and len(data) > 15:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "PLNAL":
                if data is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i)
                if not isNumeric(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i)
                if data is not None and len(data) > 2:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "VORNR":
                if data is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i)
                if data is not None and len(data) > 4:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i)
                if not isNumOnly(real_data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "SUMLIMIT":
            	if data is not None and not isNumOnly(data):
            		writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i)
            	if data is None and get_value_by_row_colname(data_ws, "SUMNOLIM", i) is None:
            		writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Overall Limit Conflict with Unlimited Indicator"), i)	
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "COMMITMENT":
            	overall = get_value_by_row_colname(data_ws,"SUMLIMIT",i)
            	if data is not None and not isNumOnly(data):
            		writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i)
            	elif overall is not None and isNumOnly(overall) and data is not None:
            		if int(data) > int(overall):
            			writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Overall Limit Conflit with Expected Value"), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "WAERS":
            	if data != "THB":
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "THB"), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "SUMNOLIM":
            	if data is not None and data != "X":
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "X"), i)
                if get_value_by_row_colname(data_ws, "SUMLIMIT", i) is not None:
                    if data is not None and data != "X":
                        writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "X"), i)
                    elif data is not None and get_value_by_row_colname(data_ws, "SUMLIMIT", i) is not None:
                    	writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.UNDEFINED[1].format("Overall Limit Conflict with Unlimited Indicator"), i)
            # else:
            #     writeHeaderReport(active_ws, "", report_data, "Success")