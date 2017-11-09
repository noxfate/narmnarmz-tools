from common import *
from openpyxl.utils import get_column_letter
import openpyxl
from validateTL import ValidateError, find_in_dict

def writeHeaderReport(ws, status, data, errorMsg, debug=None):
    new_row = []
    new_row.append(status)
    if len(data) != 4:
        raise ValueError("[4.1 Task List Planned Service] Data size is not correct")
    new_row += data
    new_row.append(errorMsg)
    new_row.append(debug)
    insert_new_row(ws, new_row)

def get_value_by_row_colname(ws, colname, row):
    MIC_HEADER = 7
    col = findColumnLetterByColNameAndStartRow(ws, colname, MIC_HEADER)
    return ws[col + str(row)].value  
'''
def check_same_MatAssign_MEINS_by_meinh(dataWb, keyDict, meinh_data):
    mat_ws = dataWb.get_sheet_by_name("04 - Mat. Assign")
    keys = dict(keyDict)
    keys.pop("VORNR")
    found = find_by_keys(mat_ws, 2, 2, keys)
    found = list(found)
    found.sort()
    row = found[0] if len(found) >= 1 else 0
    if row == 0:
        return False
    MEINS_col = findColumnLetterByColNameAndStartRow(mat_ws, "MEINS", 2)
    MEINS = mat_ws[MEINS_col + str(row)].value
    if meinh_data == MEINS:
        return True
    return False  

def check_same_header_by_werks(dataWb, keyDict, data):
    ws = dataWb.get_sheet_by_name("01 - Header")
    keys = dict(keyDict)
    keys.pop("VORNR")
    found = find_by_keys(ws, 2, 2, keys)
    found = list(found)
    found.sort()
    row = found[0] if len(found) >= 1 else 0
    if row == 0:
        return False
    col = findColumnLetterByColNameAndStartRow(ws, "WERKS", 2)
    WERKS = ws[col + str(row)].value
    if data == WERKS:
        return True
    return False
'''

def validate(wb, dataWb):
    ## CONFIG HERE NA N'Narm ##
    DATA_TAB_NAME = "Task List Planned Servie" # sheet name to find data
    DATA_ROW_COUNT = 6 # how many row to skip in header
    DATA_FIELD_ROW = 1
    DATA_HEADER_ROW = 6 # what row to find by field
    ROW_START = 2 # row to start writing data
    IS_FREEZE = True # wanna freeze header ?

    active_ws = wb.get_sheet_by_name("4.1 Task List Planned Service")
    if (IS_FREEZE):
        active_ws.freeze_panes = "A"+ str(ROW_START)
    data_ws = dataWb.get_sheet_by_name(DATA_TAB_NAME)
    n_of_data = data_ws.max_row - DATA_ROW_COUNT

    # CHECK Addtional Condition 1-2
    PLNNR_col = findColumnLetterByColNameAndStartRow(data_ws, "PLNNR", DATA_HEADER_ROW)
    PLNAL_col = findColumnLetterByColNameAndStartRow(data_ws, "PLNAL", DATA_HEADER_ROW)
    VORNR_col = findColumnLetterByColNameAndStartRow(data_ws, "VORNR", DATA_HEADER_ROW)
    KTEXT_col = findColumnLetterByColNameAndStartRow(data_ws, "KTEXT1", DATA_HEADER_ROW)
    EXTRO_col = findColumnLetterByColNameAndStartRow(data_ws, "EXTROW", DATA_HEADER_ROW)
    
    for i in range(DATA_ROW_COUNT+1, n_of_data + DATA_ROW_COUNT+1):
        PLNNR = data_ws[PLNNR_col + str(i)].value
        PLNAL = data_ws[PLNAL_col + str(i)].value
        VORNR = data_ws[VORNR_col + str(i)].value
        EXTROW = data_ws[EXTRO_col + str(i)].value

        d = dict()
        d["PLNNR"] = PLNNR
        d["PLNAL"] = PLNAL
        d["VORNR"] = VORNR
        d["EXTROW"] = EXTROW
        match_cond_1 = find_by_keys(data_ws, DATA_HEADER_ROW, DATA_ROW_COUNT, d)
        # print("Cond1", match_cond_1)
        '''

        header_ws = dataWb.get_sheet_by_name("Task List Header")
        d = dict()
        d["PLNNR"] = PLNNR
        d["PLNAL"] = PLNAL
        match_cond_2 = find_by_keys(header_ws, 7, 8, d)
        # print("Cond2", match_cond_2)
        '''
        #KTEXT = data_ws[KTEXT_col + str(i)].value
        if len(match_cond_1) > 1:
            data = [PLNNR, PLNAL, VORNR, EXTROW]
            writeHeaderReport(active_ws, "ERROR", data, ValidateError.DUPLICATE_KEY[1], "N="+str(len(match_cond_1)))
        #if len(match_cond_2) < 1:
            #data = [PLNNR, PLNAL, VORNR, KTEXT]
            #writeHeaderReport(active_ws, "ERROR", data, ValidateError.UNDEFINED[1].format("Group not mapping with 1. Task List Header "), "N="+str(len(match_cond_2)))
    
    print("Fin Additional Condition")

    # Check By Field
    key = ["PLNNR", "PLNAL", "VORNR", "EXTROW"]
    for i in range(DATA_ROW_COUNT+1, n_of_data + DATA_ROW_COUNT+1):
        for j in range(1, data_ws.max_column +1):
            report_data = [
                data_ws[PLNNR_col+str(i)].value, 
                data_ws[PLNAL_col+str(i)].value,
                data_ws[VORNR_col+str(i)].value,
                data_ws[EXTRO_col+str(i)].value
            ]
            key_value = [
                data_ws[PLNNR_col+str(i)].value, 
                data_ws[PLNAL_col+str(i)].value,
                data_ws[VORNR_col+str(i)].value,
                data_ws[EXTRO_col+str(i)].value
            ]
            key_data_dict = dict(itertools.izip(key,key_value))
            field_descr = data_ws.cell(row=DATA_FIELD_ROW, column=j).value

            real_data = data_ws.cell(row=i, column=j).value
            if isinstance(real_data, int) or isinstance(real_data, long) or isinstance(real_data, float):
                data = str(real_data)
            else:
                data = real_data
                
            if data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "PLNNR":                
                if data is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i)
                if data is not None and len(data) > 15:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "PLNAL":
                if data is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i)
                if not isNumeric(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i)
                if data is not None and len(data) > 2:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "VORNR":
                if data is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i)
                if data is not None and len(data) > 4:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i)
                if not isNumOnly(real_data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "EXTROW":
                if data is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i)
                if data is not None and len(data) > 4:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i)
                if not isNumOnly(real_data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "ASNUM":
                if data is not None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NULL[1].format(field_descr), i)  
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "KTEXT1":
                if data is None:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.NOT_NULL[1].format(field_descr), i)
                if data is not None and len(data) > 40:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "MENGE":
                if data != "1":
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "1"), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "MEINS":
                if data != "AU":
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "AU"), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "TBTWR":
                if data is not None and len(data) > 11:
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.LENGTH[1].format(field_descr), i)
                if data is not None and not isNumOnly(data):
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.VALUE_TYPE[1].format(field_descr), i)
            elif data_ws.cell(row=DATA_HEADER_ROW, column=j).value == "WAERS":
                if data != "THB":
                    writeHeaderReport(active_ws, "ERROR", report_data, ValidateError.FIXED_VALUE[1].format(field_descr, "THB"), i)
            # else:
            #     writeHeaderReport(active_ws, "", report_data, "Success")