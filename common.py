import Tkinter
import tkFileDialog
import tkMessageBox
import os
import openpyxl
import sys
import itertools
from openpyxl.utils import get_column_letter, column_index_from_string
from NarmError import *

sys.dont_write_bytecode = True

cur_path = os.path.dirname(__file__)
new_unit_path = os.path.join(cur_path, 'resource','Dict', 'unit.xlsx')

def openDialog():
    root = Tkinter.Tk()
    root.withdraw()
    fileName = tkFileDialog.askopenfilename()
    print(fileName)
    return fileName

def openExcelFile(filePath):
    os.chdir(os.path.dirname(filePath))
    wb = openpyxl.load_workbook(os.path.basename(filePath))
    return wb

unit_wb = openExcelFile(new_unit_path)

def findColumnLetterByColNameAndStartRow(worksheet, value, rowNumber):
    if isinstance(worksheet.max_column, str) and worksheet.title == "Unit":
        max_col = 6
    elif isinstance(worksheet.max_column, str) and worksheet.title == "Sampling":
        max_col = 2
    else:
        max_col = worksheet.max_column

    for i in range(1, max_col+1):
        text = worksheet[get_column_letter(i)+str(rowNumber)].value
        if (text == value):
            return get_column_letter(i)
    return None

def findCellInColumnByValue(worksheet, col, value, headerRow):
    if isinstance(col, str):
        col_letter = findColumnLetterByColNameAndStartRow(worksheet, col, headerRow)
        col_n = column_index_from_string(col_letter)
    elif isinstance(col, int):
        col_n = col
    else:
        col_n = 1

    if value is None:
        return None
    for i in range(1, worksheet.max_row+1):
        cell_val = worksheet.cell(row=i, column=col_n).value
        if (cell_val is not None) and (compareCellValue(cell_val, value)):
            return worksheet.cell(row=i, column=col_n)
    return None

def findCellListInColumnByValue(worksheet, col, value, headerRow):
    if isinstance(col, str):
        col_letter = findColumnLetterByColNameAndStartRow(worksheet, col, headerRow)
    elif isinstance(col, int):
        col_letter = get_column_letter(col)

    if value is None:
        return None
    result = set()
    rowN = worksheet.max_row
    for i in range(1, rowN+1):
        cell_val = worksheet[col_letter+str(i)].value
        if (cell_val is not None) and (compareCellValue(cell_val, value)):
            result.add(worksheet[col_letter+str(i)])
    return result

def transformUnit(input):
    global unit_wb
    ws = unit_wb.get_sheet_by_name("Unit")
    technical_cell = findCellInColumnByValue(ws, "Technical", input, 1)
    if technical_cell is None:
        # raise UnitConversionError("UnitConversion Error: ", ws, "Technical", input)
        return None
    commercial_col_letter = findColumnLetterByColNameAndStartRow(ws, "Commercial", 1)
    return ws[commercial_col_letter+str(technical_cell.row)]

def transformUnitSampling(input):
    global unit_wb
    ws = unit_wb.get_sheet_by_name("Sampling")
    old_cell = findCellInColumnByValue(ws, "Old", input, 1)
    if old_cell is None:
        return None
    new_cell_letter = findColumnLetterByColNameAndStartRow(ws, "New", 1)
    return ws[new_cell_letter+str(old_cell.row)]

def insert_new_row(ws, row_data):
    n = ws.max_row
    print(n, row_data)
    new_row = n+1
    for i in range(1, len(row_data)+1):
        ws[get_column_letter(i)+str(new_row)].value = row_data[i-1]

def isNumeric(input):
    if input is None:
        return False
    return input.replace('-', '', 1).replace('.' ,'' , 1).isdigit()

def isNumOnly(input):
    if input is None:
        return False
    data = str(input).strip()
    for i in data:
        if not i.isdigit():
            return False
    return True
    
def find_by_keys(ws, headerRow, dataRowStart, keyDict):
    DATA_ROW_COUNT = dataRowStart # how many row to skip in header
    DATA_HEADER_ROW = headerRow # what row to find by field

    found = []
    for k in keyDict.keys():
        cells = findCellListInColumnByValue(ws, k, keyDict[k], DATA_HEADER_ROW)        
        rows = []
        if cells is not None:
            cells_list = list(cells)
            for i in range(len(cells_list)):
                rows.append(cells_list[i].row)
        found.append(set(rows))
    result = set.intersection(*found)
    return result

def check_duplicate_key(ws, headerRow, dataRowStart, keyDict):
    """
    return
            True: if keys are duplicate
            False: if keys are not duplicate
    """
    DATA_ROW_COUNT = dataRowStart
    DATA_HEADER_ROW = headerRow

    key_col_dict = dict()
    for k in keyDict.keys():
        key_col_dict[k] = findColumnLetterByColNameAndStartRow(ws, k, DATA_HEADER_ROW)

    found_final = []
    i = 1
    while i < ws.max_row + 1 and len(found_final) <= 1:
        found = []
        for k in keyDict.keys():
            col_n = column_index_from_string(key_col_dict[k])
            key_val = ws.cell(row=i, column=col_n).value
            if (key_val is not None) and (compareCellValue(key_val, keyDict[k])):
                found.append(i)
            else:
                break
        if len(found) == len(keyDict):
            found_final.append(i)
        i = i+1

    return len(found_final) > 1


def is_key_exist(ws, headerRow, dataRowStart, keyDict):
    """
    return
            True: if keys are exist in {ws}
            False: if keys are not exist in {ws}
    """
    DATA_ROW_COUNT = dataRowStart
    DATA_HEADER_ROW = headerRow

    key_col_dict = dict()
    for k in keyDict.keys():
        key_col_dict[k] = findColumnLetterByColNameAndStartRow(ws, k, DATA_HEADER_ROW)

    found_final = []
    i = 1
    while i < ws.max_row + 1 and len(found_final) <= 1:
        found = []
        for k in keyDict.keys():
            col_n = column_index_from_string(key_col_dict[k])
            key_val = ws.cell(row=i, column=col_n).value
            if (key_val is not None) and (compareCellValue(key_val, keyDict[k])):
                found.append(i)
            else:
                break
        if len(found) == len(keyDict):
            found_final.append(i)
            return True
        i = i + 1
    return False


def calAfterPoint(x):
    if not '.' in x:
        return 0
    return len(x) - x.index('.') - 1

def checkDecimalPlace(num, data):
    """ checkDecimalPlace(int Number of decimal Place, string data) """
    if data == "" or data is None:
        return True
    else:
        d = calAfterPoint(data)
        num = int(num)
        if abs(d) <= abs(num):
            return True
        else:
            return False

def compareCellValue(x, y):
    if isinstance(x, type(y)) or isinstance(y, type(x)):
        return x == y
    else:
        return str(x).strip() == str(y).strip()

def isNull(x):
    return x is None or x == ''
